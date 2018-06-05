# -*- coding: utf-8 -*-

import os
import sys
import openpyxl
import traceback
import time
import multiprocessing
import csv


class Main:
    # 初始化
    def __init__(self):
        # 获取程序所在目录
        self.main_path = os.path.split(os.path.abspath(sys.argv[0]))[0]

        # 建立公共变量
        self.base_data = {
            'table_mapping': {
                'ui_inter': {},
                'inter_ui': {},
            },
            'monitoring_object_and_mapping': {
                'sheet_head': [],
                'sheet_data': {},
            },
            'data_table_head': {},
            'check_result_head': {},
        }

    # 获取基础数据函数.
    def get_base_data(self):
        path_base_data = os.path.join(self.main_path, 'Base_Data.xlsx')
        f_base_data_wb = openpyxl.load_workbook(path_base_data, read_only=True)
        for temp_sheet_name in f_base_data_wb.sheetnames:
            temp_f_base_data_wb_sheet = f_base_data_wb[temp_sheet_name]
            # 读取表映射关系
            if temp_sheet_name == 'table_mapping':
                for temp_row in temp_f_base_data_wb_sheet.iter_rows():
                    self.base_data['table_mapping']['ui_inter'][temp_row[0].value] = temp_row[1].value
                    self.base_data['table_mapping']['inter_ui'][temp_row[1].value] = temp_row[0].value
            # 读取监控对象及规则
            elif temp_sheet_name == 'monitoring_object_and_mapping':
                temp_head_identify = 0
                for temp_row in temp_f_base_data_wb_sheet.iter_rows():
                    if temp_head_identify == 1:
                        # 按序获取行
                        temp_value = [i.value for i in temp_row]
                        # 从表头中获取关键字段索引号
                        active_monitor = temp_value[self.base_data['monitoring_object_and_mapping']['sheet_head'].index(
                            '监控部署状态')]
                        monitor_target = temp_value[self.base_data['monitoring_object_and_mapping']['sheet_head'].index(
                            '字段变量标识（M）')]
                        monitor_target_source = temp_value[self.base_data['monitoring_object_and_mapping'][
                            'sheet_head'].index('字段变量标识（源）')]
                        target_table_name = temp_value[self.base_data['monitoring_object_and_mapping'][
                            'sheet_head'].index('数据库源表名称（源）')]
                        monitor_sign_of_operation = temp_value[self.base_data['monitoring_object_and_mapping'][
                            'sheet_head'].index('运算符号')]
                        threshold_upper = temp_value[self.base_data['monitoring_object_and_mapping'][
                            'sheet_head'].index('上门限')]
                        threshold_lower = temp_value[self.base_data['monitoring_object_and_mapping'][
                            'sheet_head'].index('下门限')]
                        # 判断是否启用监控，如果是则继续获取关键字段
                        if active_monitor == 1:
                            if target_table_name not in self.base_data['monitoring_object_and_mapping']['sheet_data']:
                                self.base_data['monitoring_object_and_mapping']['sheet_data'][target_table_name] = {}
                            self.base_data['monitoring_object_and_mapping']['sheet_data'][target_table_name][
                                monitor_target] = [monitor_target_source,
                                                   monitor_sign_of_operation,
                                                   threshold_upper,
                                                   threshold_lower
                                                   ]
                    elif temp_head_identify == 0:
                        self.base_data['monitoring_object_and_mapping']['sheet_head'] = [i.value for i in temp_row]
                        temp_head_identify = 1
            # 读取检查结果表头保留
            elif temp_sheet_name == 'check_result_head':
                for temp_row in temp_f_base_data_wb_sheet.iter_rows():
                    self.base_data['check_result_head'][temp_row[0].value] = []
                    for temp_cell in temp_row[1:]:
                        if temp_cell.value is not None:
                            self.base_data['check_result_head'][temp_row[0].value].append(temp_cell.value)

            else:
                # 获取核查表表头
                # temp_value_1 是原始表的名字
                # temp_value_2 是数据库导出后后重新命名的名字
                for i in temp_f_base_data_wb_sheet.iter_rows(min_row=1, max_row=1):
                    temp_value_1 = [j.value for j in i]
                for i in temp_f_base_data_wb_sheet.iter_rows(min_row=2, max_row=2):
                    temp_value_2 = [j.value for j in i]

                self.base_data['data_table_head'][temp_sheet_name] = {
                    'sheet_head_sql': temp_value_2,
                    'sheet_head_target': temp_value_1,
                    'map_sheet_head': dict(zip(temp_value_2, temp_value_1))
                }

                # self.base_data['data_table_head'][temp_sheet_name]

        print(self.base_data)
        print(self.base_data['check_result_head'])

    def get_file_list(self):
        target_path = os.path.join(self.main_path,'target_data')
        target_file_list = {}
        for temp_file_name in os.listdir(target_path):
            temp_file_full_path = os.path.join(target_path, temp_file_name)
            # 判断是否是文件（而非文件夹）
            if os.path.isfile(temp_file_full_path):
                # 判断改文件是否存在监控字段
                if temp_file_name in self.base_data['monitoring_object_and_mapping']['sheet_data']:
                    target_file_list[temp_file_name] = temp_file_full_path
        # target_file_list = {i: os.path.join(target_path, i) for i in os.listdir(target_path) if os.path.isfile(
        #     os.path.join(target_path, i))}
        target_file_list_num = len(target_file_list)
        if target_file_list_num == 0:
            print('未获取到原始数据')
            time.sleep(10)
            sys.exit()
        else:
            print('已获取到原始文件 {0} 个。'.format(target_file_list_num))
            return target_file_list

    @staticmethod
    def data_loading(file_name, file_path):
        # 直接读取
        # data_list = []
        # with open(file_path, encoding='utf-8') as f:
        #     for temp_row in f.readlines():
        #         data_list.append(temp_row.split(','))
        # return data_list

        # 使用csv模块读取
        data_list = []
        with open(file_path, encoding='utf-8') as f:
            f_csv = csv.reader(temp_line.replace('\0', '') for temp_line in f)
            for temp_row in f_csv:
                data_list.append(temp_row)
        return data_list

    @staticmethod
    def check_rule(target_value, monitor_sign_of_operation, threshold_upper, threshold_lower):

        def rule_between(obj_target_value, obj_threshold_upper, obj_threshold_lower):
            try:
                temp_target_value = float(obj_target_value)
                if obj_threshold_upper <= temp_target_value <= obj_threshold_lower:
                    return 1
                else:
                    return 0
            except:
                return 2

        def greater_than(obj_target_value, obj_threshold_upper, obj_threshold_lower):
            try:
                temp_target_value = float(obj_target_value)
                if temp_target_value > obj_threshold_upper:
                    return 1
                else:
                    return 0
            except:
                return 2

        def greater_equal_than(obj_target_value, obj_threshold_upper, obj_threshold_lower):
            try:
                temp_target_value = float(obj_target_value)
                if temp_target_value >= obj_threshold_upper:
                    return 1
                else:
                    return 0
            except:
                return 2

        def less_than(obj_target_value, obj_threshold_upper, obj_threshold_lower):
            try:
                temp_target_value = float(obj_target_value)
                if temp_target_value < obj_threshold_upper:
                    return 1
                else:
                    return 0
            except:
                return 2

        def less_equal_than(obj_target_value, obj_threshold_upper, obj_threshold_lower):
            try:
                temp_target_value = float(obj_target_value)
                if temp_target_value <= obj_threshold_upper:
                    return 1
                else:
                    return 0
            except:
                return 2

        def equal_to(obj_target_value, obj_threshold_upper, obj_threshold_lower):
            try:
                temp_target_value = float(obj_target_value)
                if temp_target_value == obj_threshold_upper:
                    return 1
                else:
                    return 0
            except:
                return 2

        rule_map = {
            'between': rule_between,
            '>': greater_than,
            '>=': greater_equal_than,
            '<': less_than,
            '<=': less_equal_than,
            '=': equal_to
        }
        return rule_map[monitor_sign_of_operation](target_value, threshold_upper, threshold_lower)

    def checker(self, file_name, target_data, check_queue):
        print('this is checker')
        for temp_row in target_data:
            temp_check_result_list = [
                file_name,
                temp_row,
                {},
            ]
            for temp_monitoring_object in self.base_data['monitoring_object_and_mapping']['sheet_data'][file_name]:
                temp_monitoring_object_value = self.base_data['monitoring_object_and_mapping']['sheet_data'][file_name][
                    temp_monitoring_object]
                try:
                    monitor_sign_of_operation = temp_monitoring_object_value[1]
                    threshold_upper = temp_monitoring_object_value[2]
                    threshold_lower = temp_monitoring_object_value[3]
                    temp_monitoring_object_index = self.base_data['data_table_head'][file_name]['sheet_head_sql'].index(
                        temp_monitoring_object)
                    check_result = self.check_rule(temp_row[temp_monitoring_object_index],
                                                   monitor_sign_of_operation,
                                                   threshold_upper,
                                                   threshold_lower)
                except IndexError as e:
                    print('【错误原因】：{0}；【源表】：{1}；【数据】：{2}'.format(e, file_name, temp_row))
                    continue
                if check_result == 1:
                    continue
                # elif check_result == 0:
                #     print('【不合规字段】：{0}:{3}；【源表】：{1}；【数据】：{2}'.format(temp_monitoring_object,
                #                                                      file_name,
                #                                                      temp_row,
                #                                                      temp_row[temp_monitoring_object_index]))
                # elif check_result == 2:
                #     print('【格式非法】：{0}:{3}；【源表】：{1}；【数据】：{2}'.format(temp_monitoring_object,
                #                                                     file_name,
                #                                                     temp_row,
                #                                                     temp_row[temp_monitoring_object_index]))

                temp_check_result_list[2][temp_monitoring_object] = temp_row[temp_monitoring_object_index]
            # 非法字段不为0，则将异常字段送到queue
            if len(temp_check_result_list[2]) != 0:
                check_queue.put(temp_check_result_list)
                # print(temp_check_result_list)

    def listen_summary(self, queue):
        print('this is listen')
        summary_list = {}
        while 1:
            temp_value = queue.get()
            # print(temp_value)
            if temp_value == 'all_finish':
                return summary_list
            else:
                if temp_value[0] not in summary_list:
                    summary_list[temp_value[0]] = []
                summary_list[temp_value[0]].append(temp_value[1:])

    def child_processes(self, file_name, file_path, p_queue):
        self.checker(file_name, self.data_loading(file_name, file_path), p_queue)

    def writer(self, value_list):
        f_result_xlsx = openpyxl.Workbook()
        for temp_table in value_list:
            temp_head_list = self.base_data['check_result_head'][temp_table]
            temp_head_check_list = list(self.base_data['monitoring_object_and_mapping']['sheet_data'][
                                            temp_table].keys())
            temp_head_check_list.sort()
            if temp_table not in f_result_xlsx.sheetnames:
                f_result_xlsx.create_sheet(temp_table)
                temp_f_result_xlsx_sheet = f_result_xlsx[temp_table]
                temp_f_result_xlsx_sheet.append(temp_head_list + temp_head_check_list)

            temp_f_result_xlsx_sheet = f_result_xlsx[temp_table]
            for temp_table_row in value_list[temp_table]:
                temp_row_list = []
                for temp_column_name in temp_head_list:
                    temp_row_value = temp_table_row[0][self.base_data['data_table_head'][temp_table][
                        'sheet_head_sql'].index(temp_column_name)]
                    temp_row_list.append(temp_row_value)
                for temp_check_name in temp_head_check_list:
                    if temp_check_name in temp_table_row[1]:
                        temp_check_name_value = temp_table_row[1][temp_check_name]
                        if temp_check_name_value == '':
                            temp_row_list.append('空')
                        else:
                            temp_row_list.append(temp_table_row[1][temp_check_name])
                    else:
                        temp_row_list.append('正确')
                temp_f_result_xlsx_sheet.append(temp_row_list)
        path_base_data = os.path.join(self.main_path, 'target_data\check_result.xlsx')
        f_result_xlsx.save(path_base_data)




if __name__ == '__main__':
    star_time = time.time()
    main = Main()
    # 获取基础数据
    main.get_base_data()
    check_file_list = main.get_file_list()
    # 多进程
    process_manager = multiprocessing.Manager()
    process_queue = process_manager.Queue()
    process_pool = multiprocessing.Pool(processes=4)
    process_listen = process_pool.apply_async(main.listen_summary, args=(process_queue,))
    jobs = []
    for temp_file in check_file_list:
        job = process_pool.apply_async(main.child_processes, args=(temp_file,
                                                                   check_file_list[temp_file],
                                                                   process_queue))
        jobs.append(job)
    for job in jobs:
        job.get()
    process_queue.put('all_finish')
    process_pool.close()
    process_pool.join()
    # 获取检查结果
    check_result_list = process_listen.get()
    main.writer(check_result_list)
    print(''.join(('>>> 历时：', time.strftime('%Y/%m/%d %H:%M:%S', time.gmtime(time.time() - star_time)))))


