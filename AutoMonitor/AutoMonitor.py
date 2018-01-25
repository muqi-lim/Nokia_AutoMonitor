# -*- coding: utf-8 -*-
# __author__ = 'linxuteng'

import os
import sys
import configparser
import time
import datetime
import cx_Oracle
import subprocess
# import prettytable
# import copy
# 邮件模块
import smtplib
from email.mime.text import MIMEText
from multiprocessing.dummy import Pool as ThreadPool
import xlrd
import openpyxl
import traceback
os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'
##############################################################################
print("""
--------------------------------
    Welcome to use tools!
    Version : 1.2.0
    Author : lin_xu_teng
    E_mail : lxuteng@live.cn
--------------------------------
""")
print('\n')
exetime = int(time.strftime('%Y%m%d', time.localtime(time.time())))
if exetime > 20190101:
    print('\n')
    print('-' * 64)
    print('试用版本已过期，请联系作者！')
    print('-' * 64)
    print('\n')
    input()
    sys.exit()

print('''

update log:

2016-11-14 添加最大用户数检测通报；
2017-1-24 根据春节保障内容添加显示字段，方便监控时使用；
2017-2-8 增加部分TOP、最大用户数超过门限小区字段；
2017-2-9 新增自动关闭拥塞小区测量上报开关；
2017-2-13 监控粒度设置为 raw_monitor 时，当有符合相关条件的top小区，会在邮件主题分别使用【干扰】【拥塞】【休眠】【maxue】
            标识，已方便邮件查看;
2017-7-28 新增当同时激活多个网管指标监控时的区分标识
2017-9-24 兼容TL16A版本基站通报及监控；添加2017年下半年考核指标通报；
2017-10-16 新增零流量小区通报；
2017-10-18 优化程序config结构；
2017-11-1 raw_monitor报表中新增esrvcc切换准备失败监控报表；
2017-11-2 raw_monitor报表中新增VoLTE低接通小区监控报表；
2017-11-3 raw_monitor新增srvcc差小区、Volte低接通小区自动关闭测量功能；
2017-11-6 修复自动关闭测量功能log时间不准确问题；
2017-11-9 raw_monitor新增下行高丢包小区自动关闭测量功能；
2017-11-9 关闭测量功能时使用多进程并发进行，提高闭锁效率；
2017-11-9 设置发送邮件时段；
2017-11-13 更换 cellinfo 格式，使其不再受编码方式影响；
2017-11-14 关闭测量log增加关闭测量开始时间及结束时间；
2017-11-14 hour表volte低接通根据当天累计进行统计；
2017-11-16 volte低接通小区、srvcc切换差小区的kpi详情记录到log；
2017-11-21 增加关闭测量的enbid例外，在例外列表里面的基站不会执行自动关闭测量；
2017-11-23 休眠小区新算法，有随机接入申请但是没有RRC申请的小区；
2017-12-1 raw_monitor模块改进srvcc切换top小区关闭测量算法，支持 关闭srvcc切换开关、修改B2门限两种模式；
            关闭srvcc切换开关模式已修复当激活actGsmSrvccMeasOpt时关闭srvcc切换功能参数报错问题；
2017-12-9 raw_monitor模块增加小区可用率检测；
2017-12-9 raw_monitor模块增加volte掉话检测；
2017-12-25 raw_monitor模块支持volte掉话超过门限后调整B2门限，让其更快切换到2G；
2018-1-23 raw_monitor模块中的可以自定义启用或关闭各类监控报告；
2018-1-24 增加高流量小区监控通报；
2018-1-25 config文件结构优化；
''')

print('\n')
print('-' * 36)
print('      >>>   starting   <<<')
print('-' * 36)
print('\n\n')
time.sleep(1)


###############################################################################


class Getini:
    # 获取配置文件
    path = os.path.split(os.path.abspath(sys.argv[0]))[0]

    def __init__(self, inifile='config.ini', inipath=path):
        self.cf = configparser.ConfigParser()
        self.cf.read(''.join((inipath, '/', inifile)), encoding='utf-8-SIG')
        self.cf_sql = configparser.ConfigParser()
        self.cf_sql.read(''.join((inipath, '/', 'config_sql.ini')), encoding='utf-8-SIG')
        self.cf_email = configparser.ConfigParser()
        self.cf_email.read(''.join((inipath, '/', 'config_email.ini')), encoding='utf-8-SIG')
        self.cf_db = configparser.ConfigParser()
        self.cf_db.read(''.join((inipath, '/', 'config_db.ini')), encoding='utf-8-SIG')

    def automonitor(self):
        self.main = {}
        # 表头判断
        self.subject_overcrowding = 0
        self.subject_sleep = 0
        self.available_range = 0
        self.volte_drop = 0
        self.subject_gps = 0
        self.subject_maxue = 0
        self.subject_srvcc = 0
        self.subject_lowconnect = 0
        self.subject_dl_low_vo_loss = 0
        self.subject_highload = 0
        for h in self.cf_email.options('main'):
            self.main[h] = self.cf_email.get('main', h)

        self.actemail = self.main['actemail']

        if datetime.datetime.now().strftime('%H') not in self.main['actemail_time'].split(','):
            self.main['actemail'] = '0'
            self.actemail = self.main['actemail']

        if self.main['actemail'] == '1':
            self.email = {}
            for o in self.cf_email.options('email'):
                if o != 'receivers':
                    o_child = self.cf_email.get('email', o)
                else:
                    o_child = self.cf_email.get('email', o).split(',')
                if o_child in ('', ['']):
                    print('>>> email 设置异常，请检查！')
                    sys.exit()
                self.email[o] = o_child

        self.config = {}
        for ii in self.cf.options('main'):
            self.config[ii] = self.cf.get('main', ii)

        for i in self.cf.options('config'):
            self.config[i] = self.cf.get('config', i)

        if self.config['timeint'] == '0' or '':
            self.config['timeint'] = 1
        else:
            try:
                self.config['timeint'] = int(self.config['timeint'])
            except:
                print('>>> timeint 必须为数字，请检查！')
                sys.exit()

        if self.config['timetype'] == 'day':
            self.starttime = (datetime.date.today() - datetime.timedelta(
                days=self.config['timeint'])).strftime('%Y%m%d')
            self.endtime = datetime.date.today().strftime('%Y%m%d')
            self.starttime_topn = (
                datetime.date.today() - datetime.timedelta(days=1)
            ).strftime('%Y%m%d')
            # self.endtime_topn = datetime.date.today().strftime('%Y%m%d')
            self.endtime_topn = self.starttime_topn
            self.mainsql = 'main_kpi_day'
            self.mainvosql = 'main_vokpi_day'
            self.top_volte_sql = 'top_volte_day'
            self.top_kpi_sql = 'top_kpi_day'
            self.sleepingcell_sql = 'sleepingcell'
            self.sleep_cell_16a_hour = 'sleep_cell_16a_hour'
            self.高流量小区 = '高流量小区_day'
            self.htmlname = (datetime.date.today() - datetime.timedelta(days=1)
                             ).strftime('%Y%m%d')
            self.maxue = 'maxue_day'
            if ini.actemail == '1':
                self.subject = self.email['subject'] + (
                    datetime.date.today() - datetime.timedelta(
                        days=1)).strftime('%Y%m%d')
        elif self.config['timetype'] == 'hour':
            self.starttime = (datetime.datetime.now() - datetime.timedelta(
                hours=self.config['timeint'])).strftime('%Y%m%d%H')
            self.endtime = datetime.datetime.now().strftime('%Y%m%d%H')
            self.starttime_topn = (
                datetime.datetime.now() - datetime.timedelta(hours=1)
            ).strftime('%Y%m%d%H')
            self.endtime_topn = datetime.datetime.now().strftime('%Y%m%d%H')
            self.mainsql = 'main_kpi_hour'
            self.mainvosql = 'main_vokpi_hour'
            self.top_volte_sql = 'top_volte_hour'
            self.top_kpi_sql = 'top_kpi_hour'
            self.sleepingcell_sql = 'sleepingcell'
            self.sleep_cell_16a_hour = 'sleep_cell_16a_hour'
            self.高流量小区 = '高流量小区_hour'
            self.htmlname = datetime.datetime.now().strftime('%Y%m%d%H')
            self.maxue = 'maxue_hour'
            self.cell_hour_all_lowconnect = 'cell_hour_all_lowconnect'
            if self.actemail == '1':
                self.subject = self.email['subject'] + datetime.datetime.now(
                ).strftime('%Y%m%d%H')
        elif self.config['timetype'] == 'raw' or self.config[
            'timetype'] == 'raw_monitor':
            self.starttime = (
                datetime.datetime.now() - datetime.timedelta(minutes=(
                                                                         self.config[
                                                                             'timeint'] + 1) * 15)).strftime(
                '%Y%m%d%H%M')
            self.endtime = datetime.datetime.now().strftime('%Y%m%d%H%M')
            self.starttime_topn = (
                datetime.datetime.now() - datetime.timedelta(
                    minutes=2 * 15 - 1)).strftime('%Y%m%d%H%M')
            self.endtime_topn = datetime.datetime.now().strftime('%Y%m%d%H%M')
            self.mainsql = 'main_kpi_raw'
            self.mainvosql = 'main_vokpi_raw'
            self.top_volte_sql = 'top_volte_raw'
            self.top_kpi_sql = 'top_kpi_raw'
            self.htmlname = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            self.sleepingcell_sql = 'sleepingcell_raw'
            self.sleep_cell_16a_raw = 'sleep_cell_16a_raw'
            self.sleep_cell_16a_hour = 'sleep_cell_16a_hour'
            self.maxue = 'maxue_raw'
            self.cell_raw_all_lowconnect = 'cell_raw_all_lowconnect'
            self.cell_raw_all_lowconnect_1 = 'cell_raw_all_lowconnect_1'
            self.cell_raw_all_dl_low_vo_loss = 'cell_raw_all_dl_low_vo_loss'
            self.休眠小区 = '休眠小区'
            self.可用率 = '可用率'
            self.高流量小区 = '高流量小区_raw'
            if self.actemail == '1':
                self.subject = self.email['subject'] + datetime.datetime.now(
                ).strftime('%Y%m%d%H%M')
        else:
            print(">>> timetype 设置异常，请检查！")
        self.alarm_sql = 'alarm'
        self.config_raw_monitor_report = {}
        for o in self.cf.options('raw_monitor_report'):
            self.config_raw_monitor_report[o] = self.cf.get('raw_monitor_report', o)

        self.db = {}
        for j in self.cf_db.options('db'):
            j_child = self.cf_db.get('db', j).split(',')
            if len(j_child) != 3:
                print('>>> [db]设置异常，请检查！')
                sys.exit()
            self.db[j] = j_child

        self.SQL_name = {}
        for k in self.cf_sql.options('SQL_name'):
            k_child = self.cf_sql.get('SQL_name', k).split(',')
            if len(k_child) != 3:
                print('>>> [SQL_name] 设置异常，请检查！')
                sys.exit()
            self.SQL_name[k] = k_child

        self.sqlpath = '/'.join((self.path, 'SQL'))
        self.displaytext = ''

        self.kpi = {}
        for n in self.cf_sql.options('kpi'):
            n_child = self.cf_sql.get('kpi', n).split(',')
            if len(n_child) != 2:
                print('>>> [kpi] 设置异常，请检查！')
                sys.exit()
            self.kpi[n.lower()] = n_child

    def cellinfo(self):
        # 读取小区IP信息
        workbook = xlrd.open_workbook('/'.join((self.path, 'cellinfo.xlsx')))
        table = workbook.sheet_by_index(0)
        self.cellinfo_data = {}
        for i in range(table.nrows):
            if i == 0:
                self.cellinfo_head = table.row_values(i)
            else:
                self.cellinfo_data[table.row_values(i)[0]] = table.row_values(i)

        # 读取关闭测量例外小区
        self.Except_Enb_List = {}
        workbook = xlrd.open_workbook('/'.join((self.path, 'Except_Enb_List.xlsx')))
        for temp_table in workbook.sheet_names():
            for i in range(workbook.sheet_by_name(temp_table).nrows):
                if i > 0:
                    if temp_table not in self.Except_Enb_List:
                        self.Except_Enb_List[temp_table] = []
                    try:
                        self.Except_Enb_List[temp_table].append(str(int(workbook.sheet_by_name(
                            temp_table).row_values(i)[0])))
                    except:
                        pass


class Db:
    def __init__(self, ip, user, pwd):
        self.ip = ip
        self.user = user
        self.pwd = pwd

    # 连接数据库

    def db_connect(self):
        print('>>> loading:', self.ip, '...')
        try:
            conn = cx_Oracle.connect(self.user, self.pwd, self.ip)
            self.cc = conn.cursor()
            self.connectstate = 1
            print('>>> 连接成功!')
        except:
            print('无法连接数据库：', self.ip, ',请检查网络或数据库设置!')
            self.connectstate = 0
            traceback.print_exc()

    def getpara(self, kpi_type):
        para_sql_list = {
            'top_srvcc': 'para_srvcc',
            'top_volte_drop': 'para_srvcc',
        }
        enb_list_text_list_temp = list(dis_pm.disabledpmmeasurement_list[kpi_type].keys())
        enb_list_text_list = ''
        for i in enb_list_text_list_temp:
            enb_list_text_list += "'"
            enb_list_text_list += str(i)
            enb_list_text_list += "'"
            enb_list_text_list += ","
        enb_list_text = enb_list_text_list[:-1]
        # enb_list_text = ','.join(list(dis_pm.disabledpmmeasurement_list[kpi_type].keys()))

        if enb_list_text == '':
            enb_list_text = "''"
        sqlfullname = ''.join(
            (ini.sqlpath, '/', para_sql_list[kpi_type], '.sql'))
        sql_text_add = 'and BTS.CO_OBJECT_INSTANCE in ({0})'.format(enb_list_text)
        try:
            f_sql = open(sqlfullname)
            try:
                sql_text = f_sql.read().replace('--&1', sql_text_add)
            except:
                print('>>>  ', sqlfullname, ' 脚本异常，请检查！')
        except:
            print('>>> ', sqlfullname, ' 不存在，请检查！')
        try:
            self.dateget = self.cc.execute(sql_text)
        except:
            print('>>> 查询出错，请检查SQL脚本！！', sqlfullname)
        headlist = [child for child in self.dateget.description]
        dbdata = self.dateget.fetchall()
        return [headlist, dbdata]

    def getdata(self,
                sqlname,
                timetype='main',
                counter='default',
                threshold='0',
                top_n='11',
                codicil=''):
        threshold = str(threshold)
        print(''.join(('>>> 开始获取数据 ', ini.SQL_name[sqlname][0], ' ...')))
        sqlfullname = ''.join(
            (ini.sqlpath, '/', ini.SQL_name[sqlname][0], '.sql'))
        try:
            f = open(sqlfullname)
        except:
            print('>>> ', sqlfullname, ' 不存在，请检查！')
        try:
            if timetype == 'main':
                sqltext = f.read().replace('&1', ini.starttime).replace(
                    '&2', ini.endtime)
            elif timetype == 'top':
                sqltext = f.read().replace('&1', ini.starttime_topn).replace(
                    '&2', ini.endtime_topn)
            # top小区counter自定义
            if counter != 'default':
                sqltext = sqltext.replace('&3', counter)
            # 附加条件，指标查询细化
            if codicil != '':
                sqltext = sqltext.replace('--&6', codicil)
        except:
            print('>>>  ', sqlfullname, ' 时段设置异常，请检查！')
        # 设置top小区门限值
        try:
            sqltext = sqltext.replace('&4', threshold)
        except:
            pass
        # 设置显示top小区数
        try:
            sqltext = sqltext.replace('&5', top_n)
        except:
            pass

        try:
            self.dateget = self.cc.execute(sqltext)
        except:
            print('>>> 查询出错，请检查SQL脚本！！', sqlfullname)
            traceback.print_exc()

    def displaydata(self, datatype='main'):
        def main():
            self.kpi_dict = ini.kpi

        def top_srvcc():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_srvcc'):
                n_child = ini.cf_sql.get('top_srvcc', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def top_qci1connect():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_qci1connect'):
                n_child = ini.cf_sql.get('top_qci1connect', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def top_qci2connect():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_qci2connect'):
                n_child = ini.cf_sql.get('top_qci2connect', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def top_qci1drop():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_qci1drop'):
                n_child = ini.cf_sql.get('top_qci1drop', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def top_rrcconnect():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_rrcconnect'):
                n_child = ini.cf_sql.get('top_rrcconnect', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def top_erabconnect():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_erabconnect'):
                n_child = ini.cf_sql.get('top_erabconnect', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def top_handover():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_handover'):
                n_child = ini.cf_sql.get('top_handover', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def top_radiodrop():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_radiodrop'):
                n_child = ini.cf_sql.get('top_radiodrop', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def top_erabdrop():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_erabdrop'):
                n_child = ini.cf_sql.get('top_erabdrop', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def sleepingcell():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('sleepingcell'):
                n_child = ini.cf_sql.get('sleepingcell', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def sleep_cell_16a():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('sleep_cell_16a'):
                n_child = ini.cf_sql.get('sleep_cell_16a', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def overcrowding():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('overcrowding'):
                n_child = ini.cf_sql.get('overcrowding', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def alarm():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('alarm'):
                n_child = ini.cf_sql.get('alarm', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def maxue():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('maxue'):
                n_child = ini.cf_sql.get('maxue', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def top_volte_connect():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_volte_connect'):
                n_child = ini.cf_sql.get('top_volte_connect', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def top_volte_drop():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_volte_drop'):
                n_child = ini.cf_sql.get('top_volte_drop', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def top_volte_uldrop():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_volte_uldrop'):
                n_child = ini.cf_sql.get('top_volte_uldrop', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def top_volte_dldrop():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('top_volte_dldrop'):
                n_child = ini.cf_sql.get('top_volte_dldrop', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def 休眠小区():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('休眠小区'):
                n_child = ini.cf_sql.get('休眠小区', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def 可用率():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('可用率'):
                n_child = ini.cf_sql.get('可用率', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        def 高流量小区():
            self.kpi_dict = {}
            for n in ini.cf_sql.options('高流量小区'):
                n_child = ini.cf_sql.get('高流量小区', n).split(',')
                if len(n_child) != 2:
                    print('>>> [kpi] 设置异常，请检查！')
                    sys.exit()
                self.kpi_dict[n.lower()] = n_child

        datatypelist = {'main': main,
                        'top_srvcc': top_srvcc,
                        'top_qci1connect': top_qci1connect,
                        'top_qci2connect': top_qci2connect,
                        'top_qci1drop': top_qci1drop,
                        'top_rrcconnect': top_rrcconnect,
                        'top_erabconnect': top_erabconnect,
                        'top_handover': top_handover,
                        'top_radiodrop': top_radiodrop,
                        'top_erabdrop': top_erabdrop,
                        'sleepingcell': sleepingcell,
                        'sleep_cell_16a': sleep_cell_16a,
                        'overcrowding': overcrowding,
                        'alarm': alarm,
                        'top_volte_connect': top_volte_connect,
                        'top_volte_drop': top_volte_drop,
                        'top_volte_uldrop': top_volte_uldrop,
                        'top_volte_dldrop': top_volte_dldrop,
                        'maxue': maxue,
                        '休眠小区': 休眠小区,
                        '可用率': 可用率,
                        '高流量小区': 高流量小区,
                        }

        datatypelist[datatype]()
        self.headlist = [
            (child[0].lower(), self.dateget.description.index(child),
             self.kpi_dict[child[0].lower()])
            for child in self.dateget.description
            if child[0].lower() in self.kpi_dict
        ]
        self.headdata = [i[0] for i in self.headlist]
        self.headindex = [i[1] for i in self.headlist]
        self.kpirange = {i[1]: i[2] for i in self.headlist}
        self.head_data_index = dict(zip(self.headdata, self.headindex))
        self.dbdata = self.dateget.fetchall()


class Email:
    def loging(self):
        print('>>> 登录email...')
        try:
            self.smtpObj = smtplib.SMTP()
            self.smtpObj.connect(ini.email['mail_host'], 25)
            self.smtpObj.login(ini.email['mail_user'], ini.email['mail_pwd'])
            print('>>> 登录email成功。')
        except:
            print('>>> email登录失败，请检查！')
            traceback.print_exc()
            # sys.exit()

    def emailtext(self, text):
        self.maintext = text

    def sendemail(self):
        # self.message = MIMEText(self.maintext, 'plain', 'utf-8')
        self.message = MIMEText(self.maintext, 'html', 'utf-8')
        temp_subject = ini.email_title_sub + '-'
        if ini.config['timetype'] == 'raw_monitor':
            if ini.subject_sleep == 1:
                temp_subject += '【接入异常】'
            if ini.available_range == 1:
                temp_subject += '【可用率】'
            if ini.volte_drop == 1:
                temp_subject += '【掉话】'
            if ini.subject_srvcc == 1:
                temp_subject += '【eSRVCC】'
            if ini.subject_lowconnect == 1:
                temp_subject += '【VOLTE低接通】'
            if ini.subject_overcrowding == 1:
                temp_subject += '【拥塞】'
            if ini.subject_dl_low_vo_loss == 1:
                temp_subject += '【高丢包】'
            if ini.subject_gps == 1:
                temp_subject += '【干扰】'
            if ini.subject_maxue == 1:
                temp_subject += '【maxue】'
            if ini.subject_highload == 1:
                temp_subject += '【高流量】'
        self.message['Subject'] = temp_subject + ini.subject
        self.message['From'] = ini.email['mail_user']
        if len(ini.email['receivers']) > 1:
            self.message['To'] = ';'.join(ini.email['receivers'])
        else:
            self.message['To'] = ini.email['receivers'][0]

        self.message["Accept-Language"] = "zh-CN"
        self.message["Accept-Charset"] = "ISO-8859-1,utf-8"

        self.smtpObj.sendmail(ini.email['mail_user'], ini.email['receivers'],
                              self.message.as_string())
        print('>>> email已发送！')


class Html:
    def __init__(self):
        self.MIMEtext = ''

    def head(self):
        self.MIMEtext += '''
            <!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <title>'''
        self.MIMEtext += ini.title
        self.MIMEtext += '''</title>
            </head>
            <body>'''

    def body(self, html_type, html_text):
        self.MIMEtext += '<%s><pre>%s</pre></%s>' % (html_type, html_text,
                                                     html_type)

    def table(self):
        self.MIMEtext += '''
         <table border = '1' cellspacing="0">
             <thead align = 'center' style = "background:#F2F2F2">
        <tr>
        '''
        for i in db.headdata:
            self.MIMEtext += '<td><b> '
            self.MIMEtext += i.upper()
            self.MIMEtext += '</b></td>'
            # 添加中文小区名和ip
            if db.kpirange[db.head_data_index[i]][1] == 'cellname':
                self.MIMEtext += '<td><b>cellname</b></td><td><b>ip</b></td>'
        self.MIMEtext += '''</tr></thead><tbody align = 'center'>'''
        for j in db.dbdata:
            self.MIMEtext += '<tr>'
            for k in db.headindex:
                self.MIMEtext += '<td>'
                if r_range(j[k], db.kpirange[k]):
                    self.MIMEtext += '<font color="#ff0000"><b>'
                if str(j[k]).count('.') == 1:
                    # self.MIMEtext += str(round(float(j[k]), 3))
                    try:
                        self.MIMEtext += str(round(float(j[k]), 3))
                    except:
                        self.MIMEtext += str(j[k])
                else:
                    self.MIMEtext += str(j[k])
                if r_range(j[k], db.kpirange[k]):
                    self.MIMEtext += '</b></font>'
                self.MIMEtext += '</td>'
                # 添加中文小区名和ip
                if db.kpirange[k][1] == 'cellname':
                    self.MIMEtext += '<td>'
                    try:
                        self.MIMEtext += ini.cellinfo_data[j[k]][2]
                    except:
                        self.MIMEtext += '-'
                    self.MIMEtext += '</td><td>'
                    try:
                        self.MIMEtext += ini.cellinfo_data[j[k]][1]
                    except:
                        self.MIMEtext += '-'
                    self.MIMEtext += '</td>'

            self.MIMEtext += '</tr>'

        self.MIMEtext += '</tbody></table>'

    def foot(self):
        self.MIMEtext += '</body></html>'

    def write_top_cell(self, type_name):
        top_cell_list_path = os.path.join(ini.path, 'HTML_TEMP/Top_Cell_List.xlsx')
        if not os.path.exists(top_cell_list_path):
            workbook = openpyxl.Workbook()
            worksheet = workbook.create_sheet(type_name)
            row_head_value_list = []
            for i in db.headdata:
                row_head_value_list.append(i)
                # 添加中文小区名和ip
                if db.kpirange[db.head_data_index[i]][1] == 'cellname':
                    row_head_value_list.append('cellname')
                    row_head_value_list.append('ip')
            worksheet.append(row_head_value_list)
            workbook.save(top_cell_list_path)
        workbook = openpyxl.load_workbook(top_cell_list_path)
        if type_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(type_name)
            row_head_value_list = []
            for i in db.headdata:
                row_head_value_list.append(i)
                # 添加中文小区名和ip
                if db.kpirange[db.head_data_index[i]][1] == 'cellname':
                    row_head_value_list.append('cellname')
                    row_head_value_list.append('ip')
            worksheet.append(row_head_value_list)
        else:
            worksheet = workbook[type_name]
        row_value_list = []
        for j in db.dbdata:
            for k in db.headindex:
                row_value_list.append(j[k])
                if db.kpirange[k][1] == 'cellname':
                    try:
                        row_value_list.append(ini.cellinfo_data[j[k]][2])
                    except:
                        row_value_list.append('-')
                    try:
                        row_value_list.append(ini.cellinfo_data[j[k]][1])
                    except:
                        row_value_list.append('-')
            worksheet.append(row_value_list)
            row_value_list = []
        workbook.save(top_cell_list_path)


def r_range(value, list):
    try:
        if list[0] == '>':
            return float(value) > float(list[1])
        elif list[0] == '>=':
            return float(value) >= float(list[1])
        elif list[0] == '<':
            return float(value) < float(list[1])
        elif list[0] == '<=':
            return float(value) <= float(list[1])
        elif list[0] == '=':
            return float(value) == float(list[1])
        elif list[0] == '!=':
            return float(value) != float(list[1])
        elif list[0] == '':
            return 0
    except:
        pass


class Report:
    def __init__(self):
        # 如果改时段没有top小区，则显示 ‘无'
        self.topcelln = 0
        # 读取小区基础信息
        ini.cellinfo()

    def day_hour_report(self, type):
        # 获取常用kpi指标
        db.getdata(ini.mainsql)
        db.displaydata()
        html.head()
        html.body('h1', '一、概况')
        html.table()

        # 休眠小区
        db.getdata(ini.sleepingcell_sql)
        db.displaydata(datatype='sleepingcell')
        if len(db.dbdata) != 0:
            html.body('h2', '  休眠小区')
            html.table()

        # 休眠小区1
        db.getdata(ini.sleep_cell_16a_hour)
        db.displaydata(datatype='sleep_cell_16a')
        if len(db.dbdata) != 0:
            html.body('h2', '  休眠及零流量小区')
            html.table()

        # 最大激活用户数检测
        db.getdata(ini.maxue, timetype='top')
        db.displaydata(datatype='maxue')
        if len(db.dbdata) != 0:
            html.body('h2', '  最大激活用户数检测')

            if type == 'hour':
                html.body('h4', '   最近一小时最大激活用户数超过配置门限小区，请尽快处理！')
            elif type == 'day':
                html.body('h4', '   昨天最大激活用户数超过配置门限小区，请尽快处理！')
            elif type == 'raw':
                html.body('h4', '   最近15分钟最大激活用户数超过配置门限小区，请尽快处理！')
            html.table()
            self.topcelln += 1

        # 获取top小区指标
        if type == 'hour':
            html.body('h1', '二、最近一小时 TOP N 小区')
        elif type == 'day':
            html.body('h1', '二、昨天 TOP N 小区')
        elif type == 'raw':
            html.body('h1', '二、最近15分钟 TOP N 小区')

        # html.body('h2', '  1、VOLTE')
        # esrvcc成功率
        db.getdata(ini.top_volte_sql, timetype='top', counter='esrvcc切换失败次数ZB')
        db.displaydata(datatype='top_srvcc')
        if len(db.dbdata) != 0:
            html.body('h3', '    ◎  esrvcc成功率')
            html.table()
            self.topcelln += 1

        # QCI低接通小区
        if ini.config['timetype'] == 'hour':
            db.getdata(
                ini.cell_hour_all_lowconnect, timetype='top', counter='volte低接通小区数', top_n='100')
        else:
            db.getdata(
                ini.top_volte_sql, timetype='top', counter='volte低接通小区数', top_n='100')
        db.displaydata(datatype='top_volte_connect')
        if len(db.dbdata) != 0:
            if ini.config['timetype'] == 'hour':
                html.body('h3', '    ◎ volte低接通小区(当天累计)')
            else:
                html.body('h3', '    ◎ volte低接通小区')
            html.table()
            self.topcelln += 1

        # QCI高掉话小区
        db.getdata(
            ini.top_volte_sql, timetype='top', counter='volte高掉话小区数', top_n='100')
        db.displaydata(datatype='top_volte_drop')
        if len(db.dbdata) != 0:
            html.body('h3', '    ◎ volte高掉话小区')
            html.table()
            self.topcelln += 1

        # QCI上行高丢包小区
        db.getdata(
            ini.top_volte_sql, timetype='top', counter='volte上行高丢包小区数', top_n='100')
        db.displaydata(datatype='top_volte_uldrop')
        if len(db.dbdata) != 0:
            html.body('h3', '    ◎ volte上行高丢包小区')
            html.table()
            self.topcelln += 1

        # QCI下行高丢包小区
        db.getdata(
            ini.top_volte_sql, timetype='top', counter='volte下行高丢包小区数', top_n='100')
        db.displaydata(datatype='top_volte_dldrop')
        if len(db.dbdata) != 0:
            html.body('h3', '    ◎ volte下行高丢包小区')
            html.table()
            self.topcelln += 1

        # 高流量小区
        db.getdata(
            ini.高流量小区, timetype='top', counter='高流量小区', top_n='100')
        db.displaydata(datatype='高流量小区')
        if len(db.dbdata) != 0:
            html.body('h3', '    ◎ 高流量小区(当天累计)')
            html.table()
            self.topcelln += 1

        # qci1_erab建立成功率
        db.getdata(
            ini.top_volte_sql, timetype='top', counter='qci1_erab建立失败次数')
        db.displaydata(datatype='top_qci1connect')
        if len(db.dbdata) != 0:
            html.body('h3', '    ◎ qci1_erab建立成功率')
            html.table()
            self.topcelln += 1

        # qci2_erab建立成功率
        db.getdata(
            ini.top_volte_sql, timetype='top', counter='qci2_erab建立失败次数')
        db.displaydata(datatype='top_qci2connect')
        if len(db.dbdata) != 0:
            html.body('h3', '    ◎ qci2_erab建立成功率')
            html.table()
            self.topcelln += 1

        # QCI1掉线率
        db.getdata(ini.top_volte_sql, timetype='top', counter='QCI1掉线次数')
        db.displaydata(datatype='top_qci1drop')
        if len(db.dbdata) != 0:
            html.body('h3', '    ◎ QCI1掉线率')
            html.table()
            self.topcelln += 1

        # html.body('h2', '  2、关键 KPI')
        # RRC连接建立成功率
        db.getdata(ini.top_kpi_sql, timetype='top', counter='RRC连接建立失败次数')
        db.displaydata(datatype='top_rrcconnect')
        if len(db.dbdata) != 0:
            html.body('h3', '    ◎ RRC连接建立成功率')
            html.table()
            self.topcelln += 1

        # 切换请求次数QQ
        db.getdata(ini.top_kpi_sql, timetype='top', counter='切换失败次数QQ')
        db.displaydata(datatype='top_handover')
        if len(db.dbdata) != 0:
            html.body('h3', '    ◎ 切换成功率QQ')
            html.table()
            self.topcelln += 1

        # ERAB建立成功率
        db.getdata(ini.top_kpi_sql, timetype='top', counter='ERAB建立失败数')
        db.displaydata(datatype='top_erabconnect')
        if len(db.dbdata) != 0:
            html.body('h3', '   ◎ ERAB建立成功率')
            html.table()
            self.topcelln += 1

        # 无线掉线率
        db.getdata(ini.top_kpi_sql, timetype='top', counter='无线掉线率分子')
        db.displaydata(datatype='top_radiodrop')
        if len(db.dbdata) != 0:
            html.body('h3', '    ◎ 无线掉线率')
            html.table()
            self.topcelln += 1

        # ERAB掉线率
        db.getdata(ini.top_kpi_sql, timetype='top', counter='ERAB掉线次数')
        db.displaydata(datatype='top_erabdrop')
        if len(db.dbdata) != 0:
            html.body('h3', '   ◎  ERAB掉线率')
            html.table()
            self.topcelln += 1

        if self.topcelln == 0:
            html.body('h3', '   ◎    无')

        # html结束
        html.foot()
        # 生成HTML文件
        # f_html = open(
        #     ''.join((ini.path, '/', 'HTML_TEMP', '/', db.ip, '-', ini.htmlname, '.html')),
        #     'w',
        #     encoding='utf-8')
        f_html = open(
            ''.join((ini.path, '/', 'HTML_TEMP', '/', ini.title, '.html')),
            'w',
            encoding='utf-8')
        f_html.write(html.MIMEtext)
        f_html.close()
        print('>>> 数据获取完成1，已生成html报告!')

    def raw_monitor(self, type):
        html.head()
        html.body('h1', 'HI，最近15分钟存在KPI恶化明显小区，可能对整网指标影响较大，请尽快处理！')

        # 休眠小区，有随机接入申请，但是全部失败
        if ini.config_raw_monitor_report['active_sleep_cell_report'] == '1':
            db.getdata(ini.休眠小区, timetype='top', counter='PREAMBLE_REQ')
            db.displaydata(datatype='休眠小区')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  接入异常小区')
                html.table()
                self.topcelln += 1
                ini.subject_sleep = 1
            if len(db.dbdata) != 0:
                try:
                    html.write_top_cell('接入异常小区')
                except:
                    pass

        # 可用率异常小区
        if ini.config_raw_monitor_report['active_available_report'] == '1':
            db.getdata(ini.可用率)
            db.displaydata(datatype='可用率')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  可用率异常小区')
                html.table()
                self.topcelln += 1
                ini.available_range = 1

        # volte高掉话小区
        if ini.config_raw_monitor_report['active_volte_drop_report'] == '1':
            db.getdata(
                ini.top_kpi_sql,
                timetype='top',
                counter='QCI1掉线次数_监控_1',
                threshold=ini.config['volte_drop_num'],
                top_n='999'
            )
            db.displaydata(datatype='top_volte_drop')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  volte掉话')

                # 保留TOP小区信息
                if ini.config['autodisabledpmmeasurement'] == '1' and ini.config['enable_cell_raw_all_volte_drop'] == '1':
                    dis_pm.autodisabledpmmeasurementdata['mark'] = 1
                    dis_pm.autodisabledpmmeasurementdata['top_volte_drop'] = [temp_i[2] for temp_i in db.dbdata]
                    html.body('h3', '<font color="#ff0000"><b>     !!!注意!!! 以下小区volte掉话较多,'
                                    '已尝试激活esrvcc切换开关（actSrvccToGsm）及调整B2门限!!!</b></font>')
                    html.body('h3', '<font color="#ff0000"><b>       >>>请尽快处理并恢复测量开关！<<<</b></font>')

                html.table()
                self.topcelln += 1
                ini.volte_drop = 1
                try:
                    html.write_top_cell('volte掉话')
                except:
                    pass

        # esrvcc切换差小区
        if ini.config_raw_monitor_report['active_srvcc_report'] == '1':
            db.getdata(
                ini.top_kpi_sql,
                timetype='top',
                counter='esrvcc切换失败次数ZB',
                threshold=ini.config['top_srvcc_fail_num'],
            )
            db.displaydata(datatype='top_srvcc')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  esrvcc切换失败次数ZB')

                # 保留TOP小区信息
                if ini.config['autodisabledpmmeasurement'] == '1' and ini.config['enable_top_srvcc'] == '1':
                    dis_pm.autodisabledpmmeasurementdata['mark'] = 1
                    dis_pm.autodisabledpmmeasurementdata['top_srvcc'] = [temp_i[2] for temp_i in db.dbdata]
                    if ini.config['top_srvcc_type'] == 'b2':
                        html.body('h3', '<font color="#ff0000"><b>     !!!注意!!! 以下小区esrvcc切换失败较多,'
                                        '已尝试调整B2门限，减慢其向2G切换!!!</b></font>')
                    else:
                        html.body('h3', '<font color="#ff0000"><b>     !!!注意!!! 以下小区esrvcc切换失败较多,'
                                        '已尝试将esrvcc切换开关（actSrvccToGsm）关闭!!!</b></font>')
                    html.body('h3', '<font color="#ff0000"><b>       >>>请尽快处理并恢复！<<<</b></font>')

                html.table()
                self.topcelln += 1
                ini.subject_srvcc = 1
                try:
                    html.write_top_cell('esrvcc切换差小区')
                except:
                    pass

        # volte低接通小区,小于1erl小区监控
        if ini.config_raw_monitor_report['active_volte_connect_report'] == '1':
            db.getdata(
                ini.cell_raw_all_lowconnect, timetype='top', counter='qci1无线接通率')
            db.displaydata(datatype='top_volte_connect')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  VOLTE低接通（小于1erl小区）')

                # 保留TOP小区信息
                if ini.config['autodisabledpmmeasurement'] == '1' and ini.config['enable_top_volte_connect'] == '1':
                    dis_pm.autodisabledpmmeasurementdata['mark'] = 1
                    dis_pm.autodisabledpmmeasurementdata['top_volte_connect'] = [temp_i[2] for temp_i in db.dbdata]
                    html.body('h3', '<font color="#ff0000"><b>     !!!注意!!! 以下小区可能演变换成volte低接通小区,'
                                    '已尝试将测量开关（mtEPSBearer）关闭!!!</b></font>')
                    html.body('h3', '<font color="#ff0000"><b>       >>>请尽快处理并恢复测量开关！<<<</b></font>')

                html.table()
                self.topcelln += 1
                ini.subject_lowconnect = 1
                try:
                    html.write_top_cell('volte接通差小区')
                except:
                    pass

        # volte低接通小区，大于1erl小区监控
        if ini.config_raw_monitor_report['active_volte_connect_1_report'] == '1':
            db.getdata(
                ini.cell_raw_all_lowconnect_1, timetype='top', counter='qci1无线接通率')
            db.displaydata(datatype='top_volte_connect')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  VOLTE低接通（大于1erl小区）')

                # 保留TOP小区信息
                if ini.config['autodisabledpmmeasurement'] == '1' and ini.config[
                    'enable_top_volte_connect_1'] == '1':
                    dis_pm.autodisabledpmmeasurementdata['mark'] = 1
                    dis_pm.autodisabledpmmeasurementdata['top_volte_connect_1'] = [temp_i[2] for temp_i in
                                                                                 db.dbdata]
                    html.body('h3', '<font color="#ff0000"><b>     !!!注意!!! 以下小区可能演变换成volte低接通小区,'
                                    '已尝试将测量开关（mtUEstate）关闭!!!</b></font>')
                    html.body('h3', '<font color="#ff0000"><b>       >>>请尽快处理并恢复测量开关！<<<</b></font>')

                html.table()
                self.topcelln += 1
                ini.subject_lowconnect = 1
                try:
                    html.write_top_cell('volte接通差小区_1')
                except:
                    pass

        # 高拥塞小区
        if ini.config_raw_monitor_report['active_overcrowding_report'] == '1':
            db.getdata(
                ini.top_kpi_sql, timetype='top', counter='拥塞次数', threshold=ini.config['overcrowding_num'])
            db.displaydata(datatype='overcrowding')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  高拥塞小区')

                # 保留高拥塞小区信息
                if ini.config['autodisabledpmmeasurement'] == '1' and ini.config['enable_overcrowding'] == '1':
                    dis_pm.autodisabledpmmeasurementdata['mark'] = 1
                    dis_pm.autodisabledpmmeasurementdata['overcrowding'] = [temp_i[2] for temp_i in db.dbdata]
                    html.body('h3', '<font color="#ff0000"><b>     !!!注意!!! 以下小区因拥塞对现网指标影响较大，'
                                    '已尝试将RRC测量上报开关关闭!!!</b></font>')
                    html.body('h3', '<font color="#ff0000"><b>       >>>请尽快处理并恢复测量开关！<<<</b></font>')
                html.table()
                self.topcelln += 1
                ini.subject_overcrowding = 1
                try:
                    html.write_top_cell('高拥塞小区')
                except:
                    pass

        # QCI1下行丢包率
        if ini.config_raw_monitor_report['active_volte_packet_loss_report'] == '1':
            db.getdata(
                ini.cell_raw_all_dl_low_vo_loss, timetype='top', counter='QCI1下行丢包率')
            db.displaydata(datatype='top_volte_dldrop')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  QCI1丢包率')

                # 保留TOP小区信息
                if ini.config['autodisabledpmmeasurement'] == '1' and ini.config['enable_cell_raw_all_dl_low_vo_loss'] == '1':
                    dis_pm.autodisabledpmmeasurementdata['mark'] = 1
                    dis_pm.autodisabledpmmeasurementdata['top_volte_dldrop'] = [temp_i[2] for temp_i in db.dbdata]
                    html.body('h3', '<font color="#ff0000"><b>     !!!注意!!! 以下小区可能演变换成volte高丢包,'
                                    '已尝试将测量开关（mtQoS）关闭!!!</b></font>')
                    html.body('h3', '<font color="#ff0000"><b>       >>>请尽快处理并恢复测量开关！<<<</b></font>')

                html.table()
                self.topcelln += 1
                ini.subject_dl_low_vo_loss = 1
                try:
                    html.write_top_cell('QCI1下行丢包率')
                except:
                    pass

        # GPS故障小区
        if ini.config_raw_monitor_report['active_gps_alarm_report'] == '1':
            db.getdata(ini.alarm_sql)
            db.displaydata(datatype='alarm')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  GPS故障基站，可能造成大面积干扰，请尽快处理！')
                html.table()
                self.topcelln += 1
                ini.subject_gps = 1

        # 休眠小区
        if ini.config_raw_monitor_report['active_sleep_cell_1_report'] == '1':
            db.getdata(ini.sleepingcell_sql)
            db.displaydata(datatype='sleepingcell')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  休眠小区')
                html.table()
                self.topcelln += 1
                ini.subject_sleep = 1

        # 休眠及零流量小区
        if ini.config_raw_monitor_report['active_sleep_cell_2_report'] == '1':
            db.getdata(ini.sleep_cell_16a_raw)
            db.displaydata(datatype='sleep_cell_16a')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  休眠及零流量小区')
                html.table()
                self.topcelln += 1
                ini.subject_sleep = 1

        # 最大激活用户数检测
        if ini.config_raw_monitor_report['active_maxue_report'] == '1':
            db.getdata(ini.maxue, timetype='top')
            db.displaydata(datatype='maxue')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  最大激活用户数检测：最近一个时段最大激活用户数超过配置门限，请尽快扩容！')
                html.table()
                self.topcelln += 1
                ini.subject_maxue = 1

        # 高流量小区
        if ini.config_raw_monitor_report['active_highload_report'] == '1':
            db.getdata(
                ini.高流量小区, timetype='top', counter='高流量小区')
            db.displaydata(datatype='高流量小区')
            if len(db.dbdata) != 0:
                html.body('h2', '   ◎  高流量小区')

                # 保留TOP小区信息
                if ini.config['autodisabledpmmeasurement'] == '1' and ini.config[
                    'enable_highload'] == '1':
                    dis_pm.autodisabledpmmeasurementdata['mark'] = 1
                    dis_pm.autodisabledpmmeasurementdata['highload'] = [temp_i[2] for temp_i in db.dbdata]
                    html.body('h3', '<font color="#ff0000"><b>     !!!注意!!! 以下小区可能成为高流量小区！</b></font>')
                    html.body('h3', '<font color="#ff0000"><b>       >>>请尽快处理！<<<</b></font>')

                html.table()
                self.topcelln += 1
                ini.subject_highload = 1
                try:
                    html.write_top_cell('高流量小区')
                except:
                    pass

        # html结束
        html.foot()
        # 生成HTML文件
        if self.topcelln != 0:
            # f_html = open(
            #     ''.join((ini.path, '/', 'HTML_TEMP', '/', ini.htmlname,
            #              '_monitor.html')),
            #     'w',
            #     encoding='utf-8')
            f_html = open(
                ''.join((ini.path, '/', 'HTML_TEMP', '/', ini.title, '.html')),
                'w',
                encoding='utf-8')
            f_html.write(html.MIMEtext)
            f_html.close()
            print('>>> 数据获取完成，已生成html报告!')
        else:
            print('>>> 该时段未存在恶化小区！')


class Autodisablepm:
    def __init__(self):
        # 初始化，并指定标识，以确定字典里面是否有数据
        self.autodisabledpmmeasurementdata = {
            'mark': 0,
            'overcrowding': [],
            'top_srvcc': [],
            'top_volte_connect': [],
            'top_volte_connect_1': [],
            'top_volte_dldrop': [],
            'top_volte_drop': [],
        }
        self.para_value_list = {
            'top_srvcc': {},
            'top_volte_drop': {},
        }
        # 每个基站生成对应此xml
        self.cmd_xml_name_list = {
            'top_srvcc': {
                'up': {},
                'bu': {}
            },
            'top_volte_drop': {
                'up': {},
                'bu': {}
            }
        }

    def format_enbid_ip(self):
        # 获取高拥塞小区，并转化成 {enbid：ip} 格式
        self.disabledpmmeasurement_list = {
            'mark': 0,
            'overcrowding': {},
            'top_srvcc': {},
            'top_volte_connect': {},
            'top_volte_connect_1': {},
            'top_volte_dldrop': {},
            'top_volte_drop': {},
        }
        # 仅检查未满足调整基站列表
        self.disabledpmmeasurement_list_pop = {
            'overcrowding': [],
            'top_srvcc': [],
            'top_volte_connect': [],
            'top_volte_connect_1': [],
            'top_volte_dldrop': [],
            'top_volte_drop': [],
        }
        top_name_tran = {
            'overcrowding': '拥塞',
            'top_srvcc': 'eSRVCC切换差小区',
            'top_volte_connect': 'Volte低接通小区',
            'top_volte_connect_1': 'Volte低接通小区_1',
            'top_volte_dldrop': 'Volte高丢包',
            'top_volte_drop': 'Volte高掉话',
        }
        for temp_table in self.autodisabledpmmeasurementdata:
            if temp_table != 'mark':
                for temp_enbid in self.autodisabledpmmeasurementdata[temp_table]:
                    if str(temp_enbid[:6]) not in ini.Except_Enb_List[top_name_tran[temp_table]]:
                        try:
                            if str(ini.cellinfo_data[temp_enbid][1]).count('.') == 3:
                                if temp_enbid[:6] not in self.disabledpmmeasurement_list[temp_table]:
                                    self.disabledpmmeasurement_list[temp_table][temp_enbid[:6]] = [ini.cellinfo_data[
                                                                                                       temp_enbid][1],
                                                                                                   [temp_enbid]
                                                                                                   ]
                                else:
                                    if temp_enbid not in self.disabledpmmeasurement_list[temp_table][temp_enbid[:6]][1]:
                                        self.disabledpmmeasurement_list[temp_table][temp_enbid[:6]][1].append(temp_enbid)
                                self.disabledpmmeasurement_list['mark'] = 1
                        except:
                            print(''.join(('>>> 基础数据 cellinfo 中未存在ENBID:', temp_enbid[:6], ' ,请检查完善！')))
        # 如果所有小区都没有读取到ip，则退出
        if self.disabledpmmeasurement_list['mark'] == 0:
            return 0
        else:
            return 1

    def para_format(self, value_list, kpi_type):
        headlist, dbdata = value_list[0], value_list[1]
        headlist = [i[0] for i in headlist]
        # 生成参数列表
        for i in dbdata:
            if i[1] in self.para_value_list[kpi_type]:
                if i[2] in self.para_value_list[kpi_type][i[1]]:
                    self.para_value_list[kpi_type][i[1]][i[2]].append(dict(zip(headlist, i)))
                else:
                    self.para_value_list[kpi_type][i[1]][i[2]] = []
                    self.para_value_list[kpi_type][i[1]][i[2]].append(dict(zip(headlist, i)))
            else:
                self.para_value_list[kpi_type][i[1]] = {}
                self.para_value_list[kpi_type][i[1]][i[2]] = []
                self.para_value_list[kpi_type][i[1]][i[2]].append(dict(zip(headlist, i)))

        for temp_enbid in self.disabledpmmeasurement_list[kpi_type]:
            self.cmd_xml_name_list[kpi_type]['up'][temp_enbid] = ''.join((
                'temp_para_',
                kpi_type,
                '_',
                temp_enbid,
                '_',
                ini.htmlname,
                '.xml'
            ))
            self.cmd_xml_name_list[kpi_type]['bu'][temp_enbid] = ''.join((
                'temp_para_',
                kpi_type,
                '_',
                temp_enbid,
                '_',
                ini.htmlname,
                '_bu.xml'
            ))
        # 对应基站生成对应的参数修改xml
        if kpi_type == 'top_srvcc':
            # 生成备份文件
            for temp_enbid in self.disabledpmmeasurement_list[kpi_type]:
                temp_xml_path_name_bu = os.path.join(
                    ini.path,
                    'CommisionTool',
                    self.cmd_xml_name_list[kpi_type]['bu'][temp_enbid]
                )
                f_bu = open(temp_xml_path_name_bu, 'w')
                f_bu.write('<raml xmlns="raml21.xsd" version="2.1">\n')
                f_bu.write('<cmData id="3221225472" scope="all" type="plan">\n')

                for temp_cellid in self.para_value_list[kpi_type][temp_enbid]:
                    if self.para_value_list[kpi_type][temp_enbid][temp_cellid][0]['ACTGSMSRVCCMEASOPT'] == 1:
                        temp_object = ''.join((
                            '<managedObject class="LNCEL" version="',
                            self.para_value_list[kpi_type][temp_enbid][temp_cellid][0]['VERSION'],
                            '" distName="MRBTS-',
                            temp_enbid,
                            '/LNBTS-',
                            temp_enbid,
                            '/LNCEL-',
                            temp_cellid.split('_')[1],
                            '" operation="update">\n'
                        ))
                        f_bu.write(temp_object)
                        f_bu.write('<p name="actGsmSrvccMeasOpt">true</p>\n')
                        f_bu.write('</managedObject>\n')
                temp_object = ''.join((
                    '<managedObject class="LNBTS" version="',
                    self.para_value_list[kpi_type][temp_enbid][temp_cellid][0]['VERSION'],
                    '" distName="MRBTS-',
                    temp_enbid,
                    '/LNBTS-',
                    temp_enbid,
                    '" operation="update">\n'
                ))
                f_bu.write(temp_object)
                f_bu.write('<p name="actSrvccToGsm">true</p>\n')
                f_bu.write('</managedObject>\n')

                for temp_cellid in self.para_value_list[kpi_type][temp_enbid]:
                    for temp_id in self.para_value_list[kpi_type][temp_enbid][temp_cellid]:
                        temp_object = ''.join((
                            '<managedObject class="LNHOG" version="',
                            temp_id['VERSION'],
                            '" distName="MRBTS-',
                            temp_enbid,
                            '/LNBTS-',
                            temp_enbid,
                            '/LNCEL-',
                            temp_cellid.split('_')[1],
                            '/LNHOG-',
                            temp_id['LNHOG_ID'],
                            '" operation="update">\n'
                        ))
                        f_bu.write(temp_object)
                        if temp_id['B2THRESHOLD1GERANQCI1'] is None:
                            f_bu.write(
                                '<p name="b2Threshold1GERANQci1"></p>\n'
                            )
                        else:
                            f_bu.write(
                                '<p name="b2Threshold1GERANQci1">{0}</p>\n'.format(
                                    int(temp_id['B2THRESHOLD1GERANQCI1']) + 140
                                )
                            )
                        if temp_id['B2THRESHOLD2RSSIGERANQCI1'] is None:
                            f_bu.write(
                                '<p name="b2Threshold2RssiGERANQci1"></p>\n'
                            )
                        else:
                            f_bu.write(
                                '<p name="b2Threshold2RssiGERANQci1">{0}</p>\n'.format(
                                    int(temp_id['B2THRESHOLD2RSSIGERANQCI1']) + 110
                                )
                            )
                        f_bu.write('</managedObject>\n')
                f_bu.write('</cmData>\n')
                f_bu.write('</raml>\n')
                f_bu.close()

            # 生成调整文件
            for temp_enbid in self.disabledpmmeasurement_list[kpi_type]:
                temp_xml_path_name = os.path.join(
                    ini.path,
                    'CommisionTool',
                    self.cmd_xml_name_list[kpi_type]['up'][temp_enbid]
                )
                f = open(temp_xml_path_name, 'w')
                f.write('<raml xmlns="raml21.xsd" version="2.1">\n')
                f.write('<cmData id="3221225472" scope="all" type="plan">\n')
                if ini.config['top_srvcc_type'] == 'disactive':
                    for temp_cellid in self.para_value_list[kpi_type][temp_enbid]:
                        if self.para_value_list[kpi_type][temp_enbid][temp_cellid][0]['ACTGSMSRVCCMEASOPT'] == 1:
                            temp_object = ''.join((
                                '<managedObject class="LNCEL" version="',
                                self.para_value_list[kpi_type][temp_enbid][temp_cellid][0]['VERSION'],
                                '" distName="MRBTS-',
                                temp_enbid,
                                '/LNBTS-',
                                temp_enbid,
                                '/LNCEL-',
                                temp_cellid.split('_')[1],
                                '" operation="update">\n'
                            ))
                            f.write(temp_object)
                            f.write('<p name="actGsmSrvccMeasOpt">false</p>\n')
                            f.write('</managedObject>\n')
                    temp_object = ''.join((
                        '<managedObject class="LNBTS" version="',
                        self.para_value_list[kpi_type][temp_enbid][temp_cellid][0]['VERSION'],
                        '" distName="MRBTS-',
                        temp_enbid,
                        '/LNBTS-',
                        temp_enbid,
                        '" operation="update">\n'
                    ))
                    f.write(temp_object)
                    f.write('<p name="actSrvccToGsm">false</p>\n')
                    f.write('</managedObject>\n')
                elif ini.config['top_srvcc_type'] == 'b2':
                    for temp_cellid in self.disabledpmmeasurement_list[kpi_type][temp_enbid][1]:
                        for temp_id in self.para_value_list[kpi_type][temp_enbid][temp_cellid]:
                            if temp_id['THRESHOLD4'] >= int(ini.config['b2Threshold1GERANQci1'.lower()]):
                                b2threshold1geranqci1 = int(temp_id['THRESHOLD4']) + 1 + 140
                            else:
                                b2threshold1geranqci1 = int(ini.config['b2Threshold1GERANQci1'.lower()]) + 140
                            if temp_id['B2THRESHOLD2RSSIGERANQCI1'] is None:
                                b2threshold2rssigeranqci1 = int(ini.config['b2Threshold2RssiGERANQci1_2'.lower()]) + 110
                            else:
                                if temp_id['B2THRESHOLD2RSSIGERANQCI1'] >= int(ini.config[
                                                                                   'b2Threshold2RssiGERANQci1_1'.lower()]):
                                    b2threshold2rssigeranqci1 = int(ini.config['b2Threshold2RssiGERANQci1_2'.lower()]) + 110
                                else:
                                    b2threshold2rssigeranqci1 = int(ini.config['b2Threshold2RssiGERANQci1_1'.lower()]) + 110

                            temp_object = ''.join((
                                '<managedObject class="LNHOG" version="',
                                temp_id['VERSION'],
                                '" distName="MRBTS-',
                                temp_enbid,
                                '/LNBTS-',
                                temp_enbid,
                                '/LNCEL-',
                                temp_cellid.split('_')[1],
                                '/LNHOG-',
                                temp_id['LNHOG_ID'],
                                '" operation="update">\n'
                            ))
                            f.write(temp_object)
                            f.write('<p name="b2Threshold1GERANQci1">{0}</p>\n'.format(b2threshold1geranqci1))
                            f.write('<p name="b2Threshold2RssiGERANQci1">{0}</p>\n'.format(b2threshold2rssigeranqci1))
                            f.write('</managedObject>\n')
                f.write('</cmData>\n')
                f.write('</raml>\n')
                f.close()
        elif kpi_type == 'top_volte_drop':
            for temp_enbid in self.disabledpmmeasurement_list[kpi_type]:
                # 生成备份文件
                temp_xml_path_name_bu = os.path.join(
                    ini.path,
                    'CommisionTool',
                    self.cmd_xml_name_list[kpi_type]['bu'][temp_enbid]
                )
                f_bu = open(temp_xml_path_name_bu, 'w')
                f_bu.write('<raml xmlns="raml21.xsd" version="2.1">\n')
                f_bu.write('<cmData id="3221225472" scope="all" type="plan">\n')
                for temp_cellid in self.para_value_list[kpi_type][temp_enbid]:
                    for temp_id in self.para_value_list[kpi_type][temp_enbid][temp_cellid]:
                        temp_object = ''.join((
                            '<managedObject class="LNHOG" version="',
                            temp_id['VERSION'],
                            '" distName="MRBTS-',
                            temp_enbid,
                            '/LNBTS-',
                            temp_enbid,
                            '/LNCEL-',
                            temp_cellid.split('_')[1],
                            '/LNHOG-',
                            temp_id['LNHOG_ID'],
                            '" operation="update">\n'
                        ))
                        f_bu.write(temp_object)
                        if temp_id['B2THRESHOLD1GERANQCI1'] is None:
                            f_bu.write(
                                '<p name="b2Threshold1GERANQci1"></p>\n'
                            )
                        else:
                            f_bu.write(
                                '<p name="b2Threshold1GERANQci1">{0}</p>\n'.format(
                                    int(temp_id['B2THRESHOLD1GERANQCI1']) + 140
                                )
                            )
                        if temp_id['B2THRESHOLD2RSSIGERANQCI1'] is None:
                            f_bu.write(
                                '<p name="b2Threshold2RssiGERANQci1"></p>\n'
                            )
                        else:
                            f_bu.write(
                                '<p name="b2Threshold2RssiGERANQci1">{0}</p>\n'.format(
                                    int(temp_id['B2THRESHOLD2RSSIGERANQCI1']) + 110
                                )
                            )
                        f_bu.write('</managedObject>\n')
                f_bu.write('</cmData>\n')
                f_bu.write('</raml>\n')
                f_bu.close()

                # 生成调整文件
                temp_xml_path_name = os.path.join(
                    ini.path,
                    'CommisionTool',
                    self.cmd_xml_name_list[kpi_type]['up'][temp_enbid]
                )
                f = open(temp_xml_path_name, 'w')
                f.write('<raml xmlns="raml21.xsd" version="2.1">\n')
                f.write('<cmData id="3221225472" scope="all" type="plan">\n')
                temp_top_volte_drop_b2_mark = 0
                for temp_cellid in self.disabledpmmeasurement_list[kpi_type][temp_enbid][1]:
                    if self.para_value_list[kpi_type][temp_enbid][temp_cellid][0]['ACTSRVCCTOGSM'] == 0:
                        temp_top_volte_drop_b2_mark = 1
                        temp_object = ''.join((
                            '<managedObject class="LNBTS" version="',
                            self.para_value_list[kpi_type][temp_enbid][temp_cellid][0]['VERSION'],
                            '" distName="MRBTS-',
                            temp_enbid,
                            '/LNBTS-',
                            temp_enbid,
                            '" operation="update">\n'
                        ))
                        f.write(temp_object)
                        f.write('<p name="actSrvccToGsm">true</p>\n')
                        f.write('</managedObject>\n')
                        break
                    else:
                        temp_srvcc_b2threshold1geranqci1_mark = 0
                        temp_srvcc_b2threshold2rssigeranqci1_mark = 0
                        for temp_id in self.para_value_list[kpi_type][temp_enbid][temp_cellid]:
                            if temp_id['B2THRESHOLD1GERANQCI1'] is None:
                                b2threshold1geranqci1 = int(ini.config['drop_b2Threshold1GERANQci1'.lower()]) + 140
                                temp_srvcc_b2threshold1geranqci1_mark = 1
                            else:
                                if temp_id['B2THRESHOLD1GERANQCI1'] < int(ini.config['drop_b2Threshold1GERANQci1'.lower()]):
                                    b2threshold1geranqci1 = int(ini.config['drop_b2Threshold1GERANQci1'.lower()]) + 140
                                    temp_srvcc_b2threshold1geranqci1_mark = 1
                            if temp_id['B2THRESHOLD2RSSIGERANQCI1'] is None:
                                b2threshold2rssigeranqci1 = int(ini.config['drop_b2Threshold2RssiGERANQci1'.lower()]) + 110
                                temp_srvcc_b2threshold2rssigeranqci1_mark = 1
                            else:
                                if temp_id['B2THRESHOLD2RSSIGERANQCI1'] > int(ini.config['drop_b2Threshold2RssiGERANQci1'.lower()]):
                                    b2threshold2rssigeranqci1 = int(ini.config['drop_b2Threshold2RssiGERANQci1'.lower()]) + 110
                                    temp_srvcc_b2threshold2rssigeranqci1_mark = 1
                            if temp_srvcc_b2threshold1geranqci1_mark + temp_srvcc_b2threshold2rssigeranqci1_mark != 0:
                                temp_top_volte_drop_b2_mark = 1
                                temp_object = ''.join((
                                    '<managedObject class="LNHOG" version="',
                                    temp_id['VERSION'],
                                    '" distName="MRBTS-',
                                    temp_enbid,
                                    '/LNBTS-',
                                    temp_enbid,
                                    '/LNCEL-',
                                    temp_cellid.split('_')[1],
                                    '/LNHOG-',
                                    temp_id['LNHOG_ID'],
                                    '" operation="update">\n'
                                ))
                                f.write(temp_object)
                                if temp_srvcc_b2threshold1geranqci1_mark == 1:
                                    f.write('<p name="b2Threshold1GERANQci1">{0}</p>\n'.format(b2threshold1geranqci1))
                                if temp_srvcc_b2threshold2rssigeranqci1_mark == 1:
                                    f.write('<p name="b2Threshold2RssiGERANQci1">{0}</p>\n'.format(b2threshold2rssigeranqci1))
                                f.write('</managedObject>\n')
                f.write('</cmData>\n')
                f.write('</raml>\n')
                f.close()
                if temp_top_volte_drop_b2_mark == 0:
                    self.disabledpmmeasurement_list_pop[kpi_type].append(temp_enbid)


    def creat_bat(self):
        # 生成命令列表
        self.cmd_list = []
        # 生成BAT文件
        self.bat_file_list = {
            'overcrowding': 'disabled_mtUEstate.xml',
            'top_srvcc': {
            },
            'top_volte_connect': 'disabled_mtEPSBearer.xml',
            'top_volte_connect_1': 'disabled_mtUEstate.xml',
            'top_volte_dldrop': 'disabled_mtQoS.xml',
        }
        self.bat_file_list_bu = {
            'overcrowding': 'enabled_mtUEstate.xml',
            'top_srvcc': {
            },
            'top_volte_connect': 'enabled_mtEPSBearer.xml',
            'top_volte_connect_1': 'enabled_mtUEstate.xml',
            'top_volte_dldrop': 'enabled_mtQoS.xml',
        }
        self.bat_path = ''.join((ini.path, '/CommisionTool/temp/DisabeledPMMeasurement_', ini.htmlname, '.bat'))
        self.bat_path_bu = ''.join((ini.path, '/CommisionTool/temp/EnabeledPMMeasurement_', ini.htmlname, '.bat'))
        f_dm = open(self.bat_path, 'w')
        f_dm_bu = open(self.bat_path_bu, 'w')
        for temp_table in self.disabledpmmeasurement_list:
            if temp_table != 'mark':
                if temp_table in ['top_srvcc', 'top_volte_drop'] and self.disabledpmmeasurement_list[temp_table] != {}:
                    self.para_format(db.getpara(temp_table), temp_table)
                for temp_enbid in self.disabledpmmeasurement_list[temp_table]:
                    if temp_enbid not in self.disabledpmmeasurement_list_pop[temp_table]:
                        if temp_table in ['top_srvcc', 'top_volte_drop']:
                            temp_text = ''.join(('call commission.bat -ne ',
                                                self.disabledpmmeasurement_list[temp_table][temp_enbid][0],
                                                 ' -pw Nemuadmin:nemuuser -deltafile ',
                                                 self.cmd_xml_name_list[temp_table]['up'][temp_enbid],
                                                 ' |tee ./temp/',
                                                 temp_table,
                                                 '-',
                                                 temp_enbid,
                                                 '-',
                                                 ini.htmlname,
                                                 '.log'))
                            temp_text_bu = ''.join(('call commission.bat -ne ',
                                                    self.disabledpmmeasurement_list[temp_table][temp_enbid][0],
                                                    ' -pw Nemuadmin:nemuuser -deltafile ',
                                                    self.cmd_xml_name_list[temp_table]['bu'][temp_enbid],
                                                    ' |tee ./temp/',
                                                    temp_table,
                                                    '-',
                                                    temp_enbid,
                                                    '-',
                                                    ini.htmlname,
                                                    '_bu.log'))
                        else:
                            temp_text = ''.join(('call commission.bat -ne ',
                                                self.disabledpmmeasurement_list[temp_table][temp_enbid][0],
                                                 ' -pw Nemuadmin:nemuuser -parameterfile ',
                                                 self.bat_file_list[temp_table],
                                                 ' |tee ./temp/',
                                                 temp_table,
                                                 '-',
                                                 temp_enbid,
                                                 '.log'))
                            temp_text_bu = ''.join(('call commission.bat -ne ',
                                                    self.disabledpmmeasurement_list[temp_table][temp_enbid][0],
                                                    ' -pw Nemuadmin:nemuuser -parameterfile ',
                                                    self.bat_file_list_bu[temp_table],
                                                    ' |tee ./temp/',
                                                    temp_table,
                                                    '-',
                                                    temp_enbid,
                                                    '_bu.log'))
                        self.cmd_list.append(temp_text)
                        f_dm.write(temp_text)
                        f_dm.write('\n')
                        f_dm_bu.write(temp_text_bu)
                        f_dm_bu.write('\n')
        f_dm.close()
        f_dm_bu.close()

    def run_call(self, ii):
        return subprocess.call(ii, shell=True)

    def run_disable_pm(self):
        # 修改运行文件夹为批处理文件所在目录，并执行批处理程序；
        os.chdir(''.join((ini.path, '/CommisionTool')))
        # subprocess.call(self.bat_path)
        pool = ThreadPool()
        pool.map(self.run_call, self.cmd_list)
        pool.close()
        pool.join()

    def creat_log(self):
        # 读取批处理程序运行结果，并生成csv记录表
        top_name_tran = {
            'overcrowding': '拥塞',
            'top_srvcc': 'eSRVCC切换差小区',
            'top_volte_connect': 'Volte低接通小区',
            'top_volte_connect_1': 'Volte低接通小区_1',
            'top_volte_dldrop': 'Volte高丢包',
            'top_volte_drop': 'Volte高掉话',
        }
        f_csv = ''.join((ini.path, '/HTML_TEMP/DisabeledPMMeasurementEnbList.csv'))
        if not os.path.exists(f_csv):
            f_csv_new = open(f_csv, 'w')
            f_csv_new.write('日期,时间,类型,enbid,ip,关闭测量情况,关闭测量开始时间,关闭测量结束时间,调整xml,备份xml\n')
            f_csv_new.close()
        with open(f_csv, 'a') as f_dml:
            for temp_table in self.disabledpmmeasurement_list:
                if temp_table != 'mark':
                    for temp_enbid in self.disabledpmmeasurement_list[temp_table]:
                        f_dml.write(time.strftime('%Y/%m/%d', time.strptime(str(ini.htmlname), '%Y%m%d%H%M%S')))
                        f_dml.write(',')
                        f_dml.write(time.strftime('%Y/%m/%d %H:%M:%S', time.strptime(str(ini.htmlname),
                                                                                     '%Y%m%d%H%M%S')))
                        f_dml.write(',')
                        f_dml.write(top_name_tran[temp_table])
                        f_dml.write(',')
                        f_dml.write(str(temp_enbid))
                        f_dml.write(',')
                        f_dml.write(str(self.disabledpmmeasurement_list[temp_table][temp_enbid][0]))
                        f_dml.write(',')
                        if temp_table in ['top_srvcc', 'top_volte_drop']:
                            log_path = ''.join((ini.path,
                                                '/CommisionTool/temp/',
                                                temp_table,
                                                '-',
                                                temp_enbid,
                                                '-',
                                                ini.htmlname,
                                                '.log'))
                        else:
                            log_path = ''.join((ini.path,
                                                '/CommisionTool/temp/',
                                                temp_table,
                                                '-',
                                                temp_enbid,
                                                '.log'))
                        try:
                            with open(log_path, 'r') as f_log:
                                log_info = f_log.read()
                                if 'Successfully activated' in log_info:
                                    f_dml.write('Successfully')
                                elif ' Connection from 0.0.0.0 exists' in log_info:
                                    f_dml.write('Connection refused')
                                elif 'Cannot connect to' in log_info:
                                    f_dml.write('Connection Failed')
                                elif 'Commissioning failed (validation failure)' in log_info:
                                    f_dml.write('Commissioning failed (validation failure)')
                                elif 'Commissioning failed' in log_info:
                                    f_dml.write('Commissioning failed')
                                elif 'Maximum number of connections has exceeded' in log_info:
                                    f_dml.write('Maximum number of connections has exceeded')
                                elif 'Failed to get HW data' in log_info:
                                    f_dml.write('Failed to get HW data')
                                elif 'Failed to read parameters' in log_info:
                                    f_dml.write('Failed to read parameters')
                                elif "Managed object can't be updated. It does not exist" in log_info:
                                    f_dml.write("Managed object can't be updated. It does not exist")
                                else:
                                    f_dml.write('Other Failed')
                                f_dml.write(',')
                            # 获取开始结束时间
                            with open(log_path, 'r') as f_log:
                                for temp_line in f_log.readlines():
                                    if 'Commissioning Tool version' in temp_line:
                                        f_dml.write(temp_line[:14])
                                        f_dml.write(',')
                                    if 'Operations finished successfully' in temp_line:
                                        f_dml.write(temp_line[:14])
                            f_dml.write(',')
                        except:
                            f_dml.write(',')
                            f_dml.write(',')
                            f_dml.write(',')
                        if temp_table in ['top_srvcc', 'top_volte_drop']:
                            f_dml.write(self.cmd_xml_name_list[temp_table]['up'][temp_enbid])
                            f_dml.write(',')
                            f_dml.write(self.cmd_xml_name_list[temp_table]['bu'][temp_enbid])
                            f_dml.write('\n')
                        else:
                            f_dml.write(self.bat_file_list[temp_table])
                            f_dml.write(',')
                            f_dml.write(self.bat_file_list_bu[temp_table])
                            f_dml.write('\n')

        print('>>> 完成！请到 /HTML_TEMP/DisabeledPMMeasurementEnbList.csv 检查运行结果.')


if __name__ == '__main__':
    print(''.join((time.strftime('%Y/%m/%d %H:%M:%S', time.localtime()))))
    ini = Getini()
    ini.automonitor()
    for db_child in ini.db:
        dis_pm = Autodisablepm()
        db = Db(ini.db[db_child][0], ini.db[db_child][1], ini.db[db_child][2])
        html = Html()
        db.db_connect()
        if db.connectstate == 0:
            sys.exit()

        # 获取生成html文件名及邮件头名
        ini.title = '_'.join((db_child, ini.cf_email.get('email', 'subject'), ini.htmlname))
        ini.email_title_sub = db_child
        db_report = Report()

        report_type = {'day': db_report.day_hour_report,
                       'hour': db_report.day_hour_report,
                       'raw': db_report.day_hour_report,
                       'raw_monitor': db_report.raw_monitor}
        report_type[ini.config['timetype']](ini.config['timetype'])

        # 如果不存在恶化小区，则不发送邮件
        if ini.config['timetype'] == 'raw_monitor' and db_report.topcelln == 0:
            ini.main['actemail'] = '0'

        # 关闭测量
        if dis_pm.autodisabledpmmeasurementdata['mark'] == 1:
            if dis_pm.format_enbid_ip() == 0:
                print('>>> 未存在需关闭测量小区。')
            else:
                print('>>> 开始尝试关闭小区测量...')
                dis_pm.creat_bat()
                dis_pm.run_disable_pm()
                dis_pm.creat_log()

        # 发送邮件
        try:
            if ini.main['actemail'] == '1':
                email = Email()
                email.loging()
                email.emailtext(html.MIMEtext)
                email.sendemail()
                email.smtpObj.close()
        except:
            traceback.print_exc()
            pass

        print('\n')
        print('='*36)

    print(''.join((time.strftime('%Y/%m/%d %H:%M:%S', time.localtime()))))
