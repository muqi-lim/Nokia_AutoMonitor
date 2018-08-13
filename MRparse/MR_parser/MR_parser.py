import xml.etree.ElementTree as ET
import numpy
import configparser
import os
import sys
import datetime
import gzip
import tarfile
import time
import multiprocessing
# import copy
import csv
import traceback
import math
import shutil
import logging
import openpyxl
from geographiclib.geodesic import Geodesic
import base64
import pyDes


def copy_right():
    logging.info('\n')
    logging.info(u"""
    --------------------------------
        Welcome to use tools!
        Author : lin_xu_teng
        E_mail : lxuteng@live.cn
    --------------------------------
    """)
    logging.info('\n')
    auth_time = int(time.strftime('%Y%m%d', time.localtime(time.time())))

    self_key = 'lxtnokia'
    self_iv = 'nokialxt'
    main_path = os.path.split(os.path.abspath(sys.argv[0]))[0]
    temp_license = open(os.path.join(main_path,'license')).read()
    k = pyDes.des(self_key, mode=pyDes.CBC, IV=self_iv, pad=None, padmode=pyDes.PAD_PKCS5)
    decryptstr = str(k.decrypt(base64.b64decode(temp_license)), encoding='utf-8').split('-')
    if decryptstr[3] == 'mr_parser':
        if auth_time > int(decryptstr[2]):
            logging.info(u'试用版本已过期，请更新！')
            input()
            sys.exit()
    else:
        logging.info(u'license错误，请重新申请！')
        input()
        sys.exit()

    logging.info(u'''
    update log:

    2017-3-22 重构，改进算法，支持多进程处理，效率大幅提高；
    2017-4-18 添加过滤器，支持指定需解码基站数据，并提取对应原始数据到指定文件夹
    2017-4-19 添加 mro_rsrp 表，记录的为MR覆盖率各个RSRP等级分布情况，同mrs解码
              的MR.RSRP表；
    2017-4-20 过滤器增加时段字段，可以指定需解码的时段；
    2017-4-23 添加进度条；
    2017-4-24 MRO覆盖率基准值由140调整为141（调整后与MRS解码的覆盖率一致）；
    2017-4-24 实现按小时级汇总；
    2017-5-2 增加解码表类型 mro_ecid_hour，此表为mro_ecid的小时级表(此更新取消)；
    2017-5-4 优化启动过滤器后的运行效率；
    2017-6-10 增加解码表类型 mro_ecid_yuan ，此表与之前袁总的解码工具解码出来的
              MRO_ECIECI 格式一致；
    2017-6-19 因 mro_ecid 表格较大，取消此表的小时级解码；
              当解码 mro_ecid 时，会同时生成 mro_ecid 与 mro_ecid_yuan 两张表；
              匹配不到的邻区不 在mro_ecid_yuan 这张表上呈现；
              优化 mro_main 中RSRP值及距离的呈现方式；
    2017-7-23 mro_main表中增加移动定义的同频重叠覆盖率；
    2017-8-8 新增运行LOG；
    2017-9-5 新增MRO解码AOA表；
    2017-10-23 修复解码MRO时部分小区部分时段缺失bug；
    2018-1-18 兼容FDD MRO解码MR覆盖率
    2018-4-23 支持按邻区频点计算频点对应覆盖率（友商竞对MR覆盖率）
    2018-6-13 修复激活MDT时，MRO解码异常bug；
    2018-6-24 添加MDT解码，当mro_parse_sheet设置为mro_rsrp_mdt时，可以生成仅带经纬度的rsrp采样点及经纬度详表；
    2018-7-18 友商竞对MR覆盖率增加按 运营商 进行汇总统计；
    2018-7-25 添加MDT重叠覆盖解析，当mro_parse_sheet设置为mro_rsrp_mdt时，新增生成mro_mdt_overlap.csv；
    2018-7-30 新增统计MRO报告数量报告，当mro_parse_sheet = mro_report_num时启用此统计报告；
    2018-8-10 MDT算法更新；
    2018-8-10 新增liscense认证方式认证；


    ''')
    logging.info(u'-' * 36)
    logging.info(u'      >>>   starting   <<<')
    logging.info(u'-' * 36)
    logging.info(u'\n')
    time.sleep(1)


################################################################################
# Module multiprocessing is organized differently in Python 3.4+
try:
    # Python 3.4+
    if sys.platform.startswith('win'):
        import multiprocessing.popen_spawn_win32 as forking
    else:
        import multiprocessing.popen_fork as forking
except ImportError:
    import multiprocessing.forking as forking

if sys.platform.startswith('win'):
    # First define a modified version of Popen.
    class _Popen(forking.Popen):
        def __init__(self, *args, **kw):
            if hasattr(sys, 'frozen'):
                # We have to set original _MEIPASS2 value from sys._MEIPASS
                # to get --onefile mode working.
                os.putenv('_MEIPASS2', sys._MEIPASS)
            try:
                super(_Popen, self).__init__(*args, **kw)
            finally:
                if hasattr(sys, 'frozen'):
                    # On some platforms (e.g. AIX) 'os.unsetenv()' is not
                    # available. In those cases we cannot delete the variable
                    # but only set it to the empty string. The bootloader
                    # can handle this case.
                    if hasattr(os, 'unsetenv'):
                        os.unsetenv('_MEIPASS2')
                    else:
                        os.putenv('_MEIPASS2', '')


    # Second override 'Popen' class with our modified version.
    forking.Popen = _Popen


################################################################################


class Main:
    def __init__(self):
        """初始化"""
        copy_right()
        self.main_path = os.path.join(os.path.split(os.path.abspath(sys.argv[0]))[0], '_config')
        self.cf = configparser.ConfigParser()
        self.cf.read(''.join((self.main_path, '\\', 'config.ini')), encoding='utf-8-SIG')
        # 初始化配置列表
        self.config_main = {}
        self.config_mrs = {}
        self.config_mro = {}
        self.config_filter = {}

        self.yesterday = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m%d')

        # 获取 main 配置
        self.get_main_config()

        # 生成结果一个空字典
        self.value_lists = {'mrs': {}, 'mro': []}

        # 统计文件数
        self.all_num = {'mrs': 0, 'mro': 0, 'mre': 0}

    def get_main_config(self):

        """获取 main 配置文件"""
        for a in self.cf.options('main'):
            self.config_main[a] = self.cf.get('main', a).split(',')

        # 部分配置特殊设置
        self.config_main['parse_type'] = [j.lower() for j in self.config_main['parse_type']]
        if self.config_main['parse_type'] == ['all']:
            self.config_main['parse_type'] = ['mrs', 'mro']

        self.config_main['file_type'] = [k.lower() for k in self.config_main['file_type']]
        if self.config_main['file_type'] == ['']:
            self.config_main['file_type'] = ['gz', 'xml']

        # 当启用定时功能时，目录重定向到 原设置目录/昨天 文件夹下
        if self.config_main['timing'][0] == '1':
            self.config_main['source_path'][0] = '\\'.join((self.config_main['source_path'][0], self.yesterday))
            self.config_main['target_path'][0] = '\\'.join((self.config_main['target_path'][0], self.yesterday))

        # process设置
        if self.config_main['process'][0] == 'min':
            self.config_main['process'][0] = '2'
        elif self.config_main['process'][0] == 'mid':
            self.config_main['process'][0] = int(multiprocessing.cpu_count() / 2 + 1)
        elif self.config_main['process'][0] in ['max', '', '0']:
            self.config_main['process'][0] = int(multiprocessing.cpu_count())
        if self.config_main['process'][0] in ['1', 1]:
            self.config_main['process'][0] = '2'

        """获取 filter 配置文件"""
        for a in self.cf.options('filter'):
            self.config_filter[a] = self.cf.get('filter', a).split(',')
        self.config_filter['filter_type'] = [i.lower() for i in self.config_filter['filter_type']]
        if self.config_filter['filter_id'] == [''] and self.config_filter['filter_type'] == ['']:
            self.config_filter['active_filter'] = 0

        # 获取处理列表
        self.parse_file_list = {}
        self.get_files(self.config_main['source_path'][0])

        # 检测结果目录是否存在，如果不存在则创建
        if len(self.parse_file_list) != 0:
            if not os.path.isdir(self.config_main['target_path'][0]):
                os.makedirs(self.config_main['target_path'][0])
        # 生成log文件
        f_log_csv = open(''.join((self.config_main['target_path'][0],
                                  '/LOG_Parse_File_List.csv'
                                  )), 'w', encoding='utf-8-sig'
                         )
        f_log_csv.write('MR_Type,File_Name,Child_File_Num,Child_File_Name\n')
        f_log_csv.close()

    def get_config(self, mr_type):

        """获取配置"""

        self.temp_mrs_data = {}
        self.temp_mro_data = {}
        # self.mro_data_data = {}
        if mr_type == 'mrs':
            for l in self.cf.options('MRS'):
                self.config_mrs[l] = self.cf.get('MRS', l).split(',')
            self.mrs_parse_sheet = []
            self.mrs_head = {}
        elif mr_type == 'mro':
            # 读取基础数据
            self.mro_earfcn_pci_cellid_relate()
            for n in self.cf.options('MRO'):
                self.config_mro[n] = self.cf.get('MRO', n).split(',')
            # RSRP对应表，用于MR覆盖率
            if self.config_mro['mro_rsrp_standard'] == ['140']:
                self.mro_rsrp_list = {
                    140: 47, 139: 47, 138: 47, 137: 47, 136: 47, 135: 47, 134: 47, 133: 47, 132: 47, 131: 47, 130: 47,
                    129: 47, 128: 47, 127: 47, 126: 47, 125: 47, 124: 47, 123: 47, 122: 47, 121: 47, 120: 47, 119: 47,
                    118: 47, 117: 47, 116: 47, 115: 47, 114: 47, 113: 47, 112: 47, 111: 47, 110: 47, 109: 47, 108: 47,
                    107: 47, 106: 47, 105: 47, 104: 47, 103: 47, 102: 47, 101: 47, 100: 47, 99: 47, 98: 47, 97: 47,
                    96: 47, 95: 47, 94: 47, 93: 47, 92: 47, 91: 47, 90: 47, 89: 47, 88: 47, 87: 47, 86: 47, 85: 47,
                    84: 47, 83: 47, 82: 47, 81: 47, 80: 47, 79: 46, 78: 46, 77: 45, 76: 45, 75: 44, 74: 44, 73: 43,
                    72: 43, 71: 42, 70: 42, 69: 41, 68: 41, 67: 40, 66: 40, 65: 39, 64: 39, 63: 38, 62: 38, 61: 37,
                    60: 37, 59: 36, 58: 35, 57: 34, 56: 33, 55: 32, 54: 31, 53: 30, 52: 29, 51: 28, 50: 27, 49: 26,
                    48: 25, 47: 24, 46: 23, 45: 22, 44: 21, 43: 20, 42: 19, 41: 18, 40: 17, 39: 16, 38: 15, 37: 14,
                    36: 13, 35: 12, 34: 11, 33: 10, 32: 9, 31: 8, 30: 7, 29: 6, 28: 5, 27: 4, 26: 3, 25: 2, 24: 1,
                    23: 1, 22: 1, 21: 1, 20: 1, 19: 0, 18: 0, 17: 0, 16: 0, 15: 0, 14: 0, 13: 0, 12: 0, 11: 0, 10: 0,
                    9: 0, 8: 0, 7: 0, 6: 0, 5: 0, 4: 0, 3: 0, 2: 0, 1: 0, 0: 0
                }
            elif self.config_mro['mro_rsrp_standard'] == ['141']:
                self.mro_rsrp_list = {
                    141: 47, 140: 47, 139: 47, 138: 47, 137: 47, 136: 47, 135: 47, 134: 47, 133: 47, 132: 47, 131: 47,
                    130: 47, 129: 47, 128: 47, 127: 47, 126: 47, 125: 47, 124: 47, 123: 47, 122: 47, 121: 47, 120: 47,
                    119: 47, 118: 47, 117: 47, 116: 47, 115: 47, 114: 47, 113: 47, 112: 47, 111: 47, 110: 47, 109: 47,
                    108: 47, 107: 47, 106: 47, 105: 47, 104: 47, 103: 47, 102: 47, 101: 47, 100: 47, 99: 47, 98: 47,
                    97: 47, 96: 47, 95: 47, 94: 47, 93: 47, 92: 47, 91: 47, 90: 47, 89: 47, 88: 47, 87: 47, 86: 47,
                    85: 47, 84: 47, 83: 47, 82: 47, 81: 47, 80: 46, 79: 46, 78: 45, 77: 45, 76: 44, 75: 44, 74: 43,
                    73: 43, 72: 42, 71: 42, 70: 41, 69: 41, 68: 40, 67: 40, 66: 39, 65: 39, 64: 38, 63: 38, 62: 37,
                    61: 37, 60: 36, 59: 35, 58: 34, 57: 33, 56: 32, 55: 31, 54: 30, 53: 29, 52: 28, 51: 27, 50: 26,
                    49: 25, 48: 24, 47: 23, 46: 22, 45: 21, 44: 20, 43: 19, 42: 18, 41: 17, 40: 16, 39: 15, 38: 14,
                    37: 13, 36: 12, 35: 11, 34: 10, 33: 9, 32: 8, 31: 7, 30: 6, 29: 5, 28: 4, 27: 3, 26: 2, 25: 1,
                    24: 1, 23: 1, 22: 1, 21: 1, 20: 0, 19: 0, 18: 0, 17: 0, 16: 0, 15: 0, 14: 0, 13: 0, 12: 0, 11: 0,
                    10: 0, 9: 0, 8: 0, 7: 0, 6: 0, 5: 0, 4: 0, 3: 0, 2: 0, 1: 0, 0: 0
                }

            # 计算ecid时对应表
            self.rsrp_dir = {
                -11: 4, -10: 5, -9: 6, -8: 7, -7: 8, -6: 9, -5: 10, -4: 11, -3: 12, -2: 13, -1: 14, 0: 15, 1: 16,
                2: 17, 3: 18, 4: 19, 5: 20, 6: 21, 7: 22, 8: 23, 9: 24, 10: 25, 11: 26
            }

            if 'mro_earfcn' in self.config_mro['mro_parse_sheet']:
                path_base_data = os.path.join(self.main_path, '频点运营商对应关系.xlsx')
                f_base_data_wb = openpyxl.load_workbook(path_base_data, read_only=True)
                self.config_mro['operator_list'] = {}
                for temp_sheet_name in f_base_data_wb.sheetnames:
                    if temp_sheet_name == '频点运营商对应关系':
                        temp_f_base_data_wb_sheet = f_base_data_wb[temp_sheet_name]
                        temp_head = []
                        for temp_row in temp_f_base_data_wb_sheet.iter_rows():
                            temp_value = [j.value for j in temp_row]
                            self.config_mro['operator_list'][temp_value[0]] = temp_value[2]

    def get_files(self, path):

        """获取需处理文件"""

        if not os.path.exists(path):
            logging.info('source_path 所设置的目录不存在，请检查！')
            sys.exit()
        for root, dirs, files in os.walk(path):
            for temp_file in files:
                temp_parse_type_num = 0
                for temp_parse_type in self.config_main['parse_type']:
                    if temp_parse_type in temp_file.lower():
                        temp_parse_type_num = 1
                        temp_file_plus = [temp_file, temp_parse_type]
                        if temp_parse_type not in self.parse_file_list:
                            self.parse_file_list[temp_parse_type] = {}
                if temp_parse_type_num == 0:
                    continue
                # 判断文件类型（gz or xml）
                # 判断是否有文件后缀，如果没有，则跳过
                if '.' not in temp_file:
                    continue
                else:
                    file_type = temp_file.split('.')[-1].lower()
                    # 如果不是 file_type 设置的文件格式，则跳过
                    if file_type not in self.config_main['file_type']:
                        continue
                    # 如果有且满足file_type设置的文件格式，则加入到parse_file_list中
                    if file_type not in self.parse_file_list[temp_file_plus[1]]:
                        self.parse_file_list[temp_file_plus[1]][file_type] = []
                source_path = os.path.join(root, temp_file)
                # 2017年5月4日新增过滤器过滤优化字段
                if self.config_filter['active_filter'] == ['1']:
                    if temp_file[-6:] == 'tar.gz' and (temp_file[-13:-11] in self.config_filter['filter_hour'] or
                                                               self.config_filter['filter_hour'] == ['']):
                        self.parse_file_list[temp_file_plus[1]][file_type].append(source_path)
                else:
                    self.parse_file_list[temp_file_plus[1]][file_type].append(source_path)
        if len(self.parse_file_list) == 0:
            logging.info('未获取到源文件，请检查source_path是否设置正确或源文件是否存在！')
            sys.exit()

    def makedir(self):
        if not os.path.exists(self.config_main['target_path'][0]):
            os.makedirs(self.config_main['target_path'][0])

    # 进度条
    @staticmethod
    def progress(num_total, num_run, file_name=''):
        bar_len = 10
        hashes = '|' * int(num_run / num_total * bar_len)
        spaces = '_' * (bar_len - len(hashes))
        sys.stdout.write("\r%s %s %s %d%%  %s" % (str(num_run), hashes + spaces, str(num_total),
                                                  int(num_run / num_total * 100), file_name))
        sys.stdout.flush()

    def mro_earfcn_pci_cellid_relate(self):

        """读取基础数据，建立earfcn pci cellid 索引表"""

        # f = open(os.path.join(config_manager.main_path, 'enb_basedat.csv'), encoding='utf-8-sig')
        # basedatas = [i.strip().split(',') for i in f.readlines()]
        # print(basedatas)

        basedatas = []
        path_base_data = os.path.join(config_manager.main_path, 'enb_basedat.xlsx')
        f_base_data_wb = openpyxl.load_workbook(path_base_data, read_only=True)
        for temp_sheet_name in f_base_data_wb.sheetnames:
            if temp_sheet_name == 'enb_basedat':
                temp_f_base_data_wb_sheet = f_base_data_wb[temp_sheet_name]
                for temp_row in temp_f_base_data_wb_sheet.iter_rows(min_row=2):
                    temp_value = [str(j.value) for j in temp_row]
                    basedatas.append(temp_value)

        self.mro_earfcn_pci_cellid = {}
        self.mro_enbid_list = {}
        self.mro_enbcellid_cn_name = {}
        self.mro_ecid_lon_lat_azi = {}

        for k in basedatas:
            if k[7] not in self.mro_earfcn_pci_cellid:
                self.mro_earfcn_pci_cellid[k[7]] = {}
            if k[6] not in self.mro_earfcn_pci_cellid[k[7]]:
                self.mro_earfcn_pci_cellid[k[7]][k[6]] = []
            try:
                self.mro_earfcn_pci_cellid[k[7]][k[6]].append((k[0], k[1], float(k[4]), float(k[5])))
            except:
                pass

            if k[0] not in self.mro_enbid_list:
                try:
                    self.mro_enbid_list[k[0]] = (float(k[4]), float(k[5]))
                except:
                    pass

            if k[1] not in self.mro_enbcellid_cn_name:
                try:
                    self.mro_enbcellid_cn_name[k[1]] = k[3]
                except:
                    pass
            temp_ecid = str(int(k[1].split('_')[0])*256 + int(k[1].split('_')[1]))
            if temp_ecid not in self.mro_ecid_lon_lat_azi:
                try:
                    self.mro_ecid_lon_lat_azi[temp_ecid] = [k[4], k[5], k[8]]
                except:
                    pass

    @staticmethod
    def distance(lon_1, lat_1, lon_2, lat_2):

        """计算距离"""

        lon1, lat1, lon2, lat2 = map(math.radians, [lon_1, lat_1, lon_2, lat_2])
        dlon = lon2 - lon1
        dlat = lat2 - lat1
        a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
        c = 2 * math.asin(math.sqrt(a))
        r = 6371  # 地球平均半径，单位为公里
        return c * r * 1000

    def min_distance_cell(self, s_enbid, n_earfcn, n_pci):

        """提取距离最近的小区"""

        min_distance_list = {}
        min_cell = '-'
        min_stance = '-'
        try:
            for i in self.mro_earfcn_pci_cellid[str(int(float(n_earfcn)//1))][str(int(float(n_pci)//1))]:
                min_distance_list[self.distance(self.mro_enbid_list[s_enbid][0], self.mro_enbid_list[s_enbid][1], i[2],
                                                i[3])] = i[1]
                min_stance = min(min_distance_list)
                min_cell = min_distance_list[min_stance]
                min_stance = int(min_stance)
        except:
            # traceback.print_exc()
            pass

        return min_cell, min_stance

    def mro_main(self, mro_object, report_time, enbid):

        """统计mro_main表"""
        temp_id = 1
        ecid = int(enbid)*256 + int(mro_object.attrib['id']) % 256
        # cmcc重叠覆盖率采集数计数器
        overlap_times = 0
        temp_mro_main = []
        for value in mro_object.iter('v'):
            temp_value = list(map(float, value.text.rstrip().replace('NIL', '0').split(' ')))
            if temp_id == 1:
                ecid_earfcn_pci = '_'.join(map(str, (ecid, temp_value[7], temp_value[8])))
                temp_mro_main = ['mro_main',
                                 ecid_earfcn_pci,
                                 [temp_value[0],
                                  temp_value[1],
                                  temp_value[2],
                                  temp_value[4],
                                  temp_value[5],
                                  temp_value[6],
                                  temp_value[20],
                                  temp_value[21],
                                  1,
                                  0,
                                  0,
                                  ]
                                 ]
                # 统计MR覆盖率
                if temp_value[0] - int(self.config_mro['mro_rsrp_standard'][0]) >= int(self.config_mro['mr_lap'][0]):
                    temp_mro_main[2][9] = 1

                temp_id = 0

            if temp_value[7] == temp_value[11]:
                if temp_value[0] - int(
                        self.config_mro['mro_rsrp_standard'][0]
                ) >= int(
                        self.config_mro['cmcc_overlap_scell_rsrp'][0]
                ) and temp_value[0] - temp_value[9] <= abs(int(self.config_mro['cmcc_overlap_db'][0])):
                    overlap_times += 1

        # 统计cmcc重叠覆盖率
        if overlap_times >= int(self.config_mro['cmcc_overlap_ncell_num'][0]):
            temp_mro_main[2][10] = 1

        # 先汇总，后才传送到queue
        try:
            self.temp_mro_data[report_time][temp_mro_main[0]][temp_mro_main[1]] += numpy.array(temp_mro_main[2])
        except:
            try:
                self.temp_mro_data[report_time][temp_mro_main[0]][temp_mro_main[1]] = numpy.array(temp_mro_main[2])
            except:
                try:
                    self.temp_mro_data[report_time][temp_mro_main[0]] = {}
                    self.temp_mro_data[report_time][temp_mro_main[0]][temp_mro_main[1]] = numpy.array(temp_mro_main[2])
                except:
                    self.temp_mro_data[report_time] = {}
                    self.temp_mro_data[report_time][temp_mro_main[0]] = {}
                    self.temp_mro_data[report_time][temp_mro_main[0]][temp_mro_main[1]] = numpy.array(temp_mro_main[2])

    def mro_ecid(self, object_mro, report_time, enbid):
        for value in object_mro.iter('v'):
            temp_value = list(map(float, value.text.rstrip().replace('NIL', '0').split(' ')))
            ecid1 = int(enbid) * 256 + int(object_mro.attrib['id']) % 256
            ecid_earfcn_pci_n_earfcn_n_pci = '_'.join(map(str, (ecid1,
                                                                temp_value[7],
                                                                temp_value[8],
                                                                temp_value[11],
                                                                temp_value[12]
                                                                )
                                                          )
                                                      )
            if temp_value[11] != 0:
                temp_mro_ecid = ['mro_ecid',
                                 ecid_earfcn_pci_n_earfcn_n_pci,
                                 [temp_value[0],
                                  temp_value[2],
                                  temp_value[9], 1,
                                  0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
                                  ]
                                 ]
                srsrp_nrsrp = temp_value[9] - temp_value[0]
                if srsrp_nrsrp < -10:
                    srsrp_nrsrp = -11
                if srsrp_nrsrp > 10:
                    srsrp_nrsrp = 11
                temp_mro_ecid[2][self.rsrp_dir[srsrp_nrsrp]] = 1
                # 先汇总，后才传送到queue
                try:
                    self.temp_mro_data[report_time][temp_mro_ecid[0]][temp_mro_ecid[1]] += numpy.array(temp_mro_ecid[2])
                except:
                    try:
                        self.temp_mro_data[report_time][temp_mro_ecid[0]][temp_mro_ecid[1]] = numpy.array(
                            temp_mro_ecid[2])
                    except:
                        try:
                            self.temp_mro_data[report_time][temp_mro_ecid[0]] = {}
                            self.temp_mro_data[report_time][temp_mro_ecid[0]][temp_mro_ecid[1]] = numpy.array(
                                temp_mro_ecid[2])
                        except:
                            self.temp_mro_data[report_time] = {}
                            self.temp_mro_data[report_time][temp_mro_ecid[0]] = {}
                            self.temp_mro_data[report_time][temp_mro_ecid[0]][temp_mro_ecid[1]] = numpy.array(
                                temp_mro_ecid[2])

    def mro_rsrp(self, mro_object, report_time, enbid):
        ecid = int(enbid)*256 + int(mro_object.attrib['id']) % 256
        for value in mro_object.iter('v'):
            temp_value = value.text.rstrip().split(' ')
            temp_mro_rsrp = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                             0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
            temp_mro_rsrp[self.mro_rsrp_list[int(temp_value[0])]] = 1
            try:
                self.temp_mro_data[report_time]['mro_rsrp'][ecid] += numpy.array(temp_mro_rsrp)
            except:
                try:
                    self.temp_mro_data[report_time]['mro_rsrp'][ecid] = numpy.array(temp_mro_rsrp)
                except:
                    try:
                        self.temp_mro_data[report_time]['mro_rsrp'] = {ecid: numpy.array(temp_mro_rsrp)}
                    except:
                        self.temp_mro_data[report_time] = {}
                        self.temp_mro_data[report_time]['mro_rsrp'] = {ecid: numpy.array(temp_mro_rsrp)}
            break

    def mro_rsrp_mdt(self, mro_object, report_time, enbid):
        ecid = int(enbid)*256 + int(mro_object.attrib['id']) % 256
        mdt_overlap_list = {
            's': [],
            'n': [],
            'mdt': []
        }
        temp_counter = 0
        for value in mro_object.iter('v'):
            temp_value = value.text.rstrip().split(' ')
            if temp_counter == 0:
                try:
                    if temp_value[27] != 'NIL':
                        # 当频点相同时
                        temp_s_rsrp = int(temp_value[0])-141
                        if temp_value[7] == temp_value[11] and temp_value[7] != 'NIL':
                            if temp_s_rsrp >= int(self.config_mro['cmcc_overlap_scell_rsrp'][0]):
                                temp_counter = 1
                                mdt_overlap_list['s'] = temp_s_rsrp
                                mdt_overlap_list['mdt'] = [temp_value[27], temp_value[28], temp_value[7], str(ecid)]
                                temp_n_rsrp = int(temp_value[9])-141
                                if temp_s_rsrp - temp_n_rsrp < int(self.config_mro['cmcc_overlap_db'][0]):
                                    mdt_overlap_list['n'].append(temp_n_rsrp)
                        else:
                            temp_counter = 2

                        temp_mro_rsrp_mdt = [0] * 50
                        temp_mro_rsrp_mdt[self.mro_rsrp_list[int(temp_value[0])]] = 1

                        temp_ecid_long_lat_value = [0] * 3
                        ecid_long_lat = '_'.join(
                            (
                                str(ecid),
                                temp_value[27],
                                temp_value[28],
                                str(int(temp_value[0])-141),
                                str(int(temp_value[1])/2-19.5),
                                str(int(temp_value[6])-11),
                                str(int(temp_value[2])*78)
                            )
                        )
                        temp_ecid_long_lat_value[0] += 1
                        try:
                            temp_ue_azi = Geodesic.WGS84.Inverse(
                                float(self.mro_ecid_lon_lat_azi[str(ecid)][1]),
                                float(self.mro_ecid_lon_lat_azi[str(ecid)][0]),
                                float(temp_value[28]), float(temp_value[27])
                            )['azi2']
                            # if temp_ue_azi < 0:
                            #     temp_ue_azi = 360+temp_ue_azi
                            # print(temp_ue_azi)
                            temp_enb_azi = float(self.mro_ecid_lon_lat_azi[str(ecid)][2])
                            if temp_enb_azi > 180:
                                temp_enb_azi = temp_enb_azi - 360
                            if (temp_enb_azi > 0 and temp_ue_azi > 0) or (temp_enb_azi < 0 and temp_ue_azi < 0):
                                temp_azi_off = abs(temp_enb_azi - temp_ue_azi)
                            else:
                                temp_azi_off_1 = abs(temp_enb_azi - temp_ue_azi)
                                temp_azi_off_2 = abs(temp_enb_azi + temp_ue_azi)
                                temp_azi_off = max(temp_azi_off_1,temp_azi_off_2)
                                if temp_azi_off > 180:
                                    temp_azi_off = 360-temp_azi_off
                            if temp_azi_off > float(self.config_mro['azi_offset'][0]):
                                temp_ecid_long_lat_value[1] += 1
                                temp_mro_rsrp_mdt[48] += 1
                                temp_ecid_long_lat_value[2] += temp_ue_azi
                                # temp_mro_rsrp_mdt[49] += temp_ue_azi

                        except:
                            traceback.print_exc()
                        try:
                            self.temp_mro_data[report_time]['mro_rsrp_mdt'][ecid] += numpy.array(temp_mro_rsrp_mdt)
                        except:
                            try:
                                self.temp_mro_data[report_time]['mro_rsrp_mdt'][ecid] = numpy.array(temp_mro_rsrp_mdt)
                            except:
                                try:
                                    self.temp_mro_data[report_time]['mro_rsrp_mdt'] = {ecid: numpy.array(temp_mro_rsrp_mdt)}
                                except:
                                    self.temp_mro_data[report_time] = {}
                                    self.temp_mro_data[report_time]['mro_rsrp_mdt'] = {ecid: numpy.array(temp_mro_rsrp_mdt)}

                        try:
                            self.temp_mro_data[report_time]['mro_rsrp_mdt_details'][ecid_long_lat] += numpy.array(temp_ecid_long_lat_value)
                        except:
                            try:
                                self.temp_mro_data[report_time]['mro_rsrp_mdt_details'][ecid_long_lat] = numpy.array(temp_ecid_long_lat_value)
                            except:
                                try:
                                    self.temp_mro_data[report_time]['mro_rsrp_mdt_details'] = {ecid_long_lat:
                                                                                                   numpy.array(temp_ecid_long_lat_value)}
                                except:
                                    self.temp_mro_data[report_time] = {}
                                    self.temp_mro_data[report_time]['mro_rsrp_mdt_details'] = {ecid_long_lat:
                                                                                                   numpy.array(temp_ecid_long_lat_value)}

                except:
                    traceback.print_exc()
            elif temp_counter == 1:
                mdt_overlap_list['n'].append(int(temp_value[9]) - 141)
            elif temp_counter == 2:
                break
        if temp_counter == 1:
            overlap_num = len(mdt_overlap_list['n'])
            if overlap_num >= int(self.config_mro['cmcc_overlap_ncell_num'][0]):
                temp_head = '_'.join(mdt_overlap_list['mdt']+[str(mdt_overlap_list['s']), str(overlap_num)])
                try:
                    self.temp_mro_data[report_time]['mro_mdt_overlap'][temp_head] += 1
                except:
                    try:
                        self.temp_mro_data[report_time]['mro_mdt_overlap'][temp_head] = 1
                    except:
                        try:
                            self.temp_mro_data[report_time]['mro_mdt_overlap'] = {temp_head: 1}
                        except:
                            self.temp_mro_data[report_time] = {}
                            self.temp_mro_data[report_time]['mro_mdt_overlap'] = {temp_head: 1}

    def mro_aoa(self, mro_object, report_time, enbid):
        ecid = int(enbid)*256 + int(mro_object.attrib['id']) % 256
        for value in mro_object.iter('v'):
            temp_value = list(map(float, value.text.rstrip().replace('NIL', '0').split(' ')))
            temp_mro_aoa = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                            0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                            0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
            temp_mro_aoa[temp_value[5]//10] = 1
            try:
                self.temp_mro_data[report_time]['mro_aoa'][ecid] += numpy.array(temp_mro_aoa)
            except:
                try:
                    self.temp_mro_data[report_time]['mro_aoa'][ecid] = numpy.array(temp_mro_aoa)
                except:
                    try:
                        self.temp_mro_data[report_time]['mro_aoa'] = {ecid: numpy.array(temp_mro_aoa)}
                    except:
                        self.temp_mro_data[report_time] = {}
                        self.temp_mro_data[report_time]['mro_aoa'] = {ecid: numpy.array(temp_mro_aoa)}
            break

    def mro_report(self, mro_object, report_time, enbid):
        ecid = int(enbid) * 256 + int(mro_object.attrib['id']) % 256
        temp_value_earfcn = 0
        temp_value_num = [0, 0, 0, 0, 0]
        for value in mro_object.iter('v'):
            temp_value = value.text.rstrip().split(' ')
            if temp_value[11] != 'NIL':
                temp_value_earfcn += 1
            else:
                break
        if temp_value_earfcn > 4:
            temp_value_earfcn = 4
        temp_value_num[temp_value_earfcn] = 1
        try:
            self.temp_mro_data[report_time]['mro_report_num'][ecid] += numpy.array(temp_value_num)
        except:
            try:
                self.temp_mro_data[report_time]['mro_report_num'][ecid] = numpy.array(temp_value_num)
            except:
                try:
                    self.temp_mro_data[report_time]['mro_report_num'] = {ecid: numpy.array(temp_value_num)}
                except:
                    self.temp_mro_data[report_time] = {}
                    self.temp_mro_data[report_time]['mro_report_num'] = {ecid: numpy.array(temp_value_num)}

    # def mro_earfcn(self, object_mro, report_time, enbid):
    #     for value in object_mro.iter('v'):
    #         temp_value = list(map(float, value.text.rstrip().replace('NIL', '0').split(' ')))
    #         ecid1 = int(enbid) * 256 + int(object_mro.attrib['id']) % 256
    #         ecid_earfcn = '_'.join(map(str, (ecid1, temp_value[11])))
    #         if temp_value[11] != 0:
    #             temp_mro_earfcn = ['mro_earfcn',
    #                                ecid_earfcn,
    #                                [0] * 96
    #                                ]
    #             temp_mro_earfcn[2][self.mro_rsrp_list[int(temp_value[0])]] = 1
    #             temp_mro_earfcn[2][self.mro_rsrp_list[int(temp_value[9])]+48] = 1
    #             # 先汇总，后才传送到queue
    #             try:
    #                 self.temp_mro_data[report_time][temp_mro_earfcn[0]][temp_mro_earfcn[1]] += numpy.array(temp_mro_earfcn[2])
    #             except:
    #                 try:
    #                     self.temp_mro_data[report_time][temp_mro_earfcn[0]][temp_mro_earfcn[1]] = numpy.array(
    #                         temp_mro_earfcn[2])
    #                 except:
    #                     try:
    #                         self.temp_mro_data[report_time][temp_mro_earfcn[0]] = {}
    #                         self.temp_mro_data[report_time][temp_mro_earfcn[0]][temp_mro_earfcn[1]] = numpy.array(
    #                             temp_mro_earfcn[2])
    #                     except:
    #                         self.temp_mro_data[report_time] = {}
    #                         self.temp_mro_data[report_time][temp_mro_earfcn[0]] = {}
    #                         self.temp_mro_data[report_time][temp_mro_earfcn[0]][temp_mro_earfcn[1]] = numpy.array(
    #                             temp_mro_earfcn[2])

    def mro_earfcn(self, object_mro, report_time, enbid):
        ecid1 = int(enbid) * 256 + int(object_mro.attrib['id']) % 256
        temp_value_list = []
        temp_mro_earfcn = [0] * 96
        temp_mro_earfcn_operator = [0] * 96
        for value in object_mro.iter('v'):
            temp_value = list(map(float, value.text.rstrip().replace('NIL', '0').split(' ')))
            if temp_value[11] != 0:
                temp_mro_earfcn[self.mro_rsrp_list[int(temp_value[9])]+48] += 1
                temp_value_list.append(int(temp_value[9]))
            else:
                break

        if len(temp_value_list) != 0:
            temp_mro_earfcn_operator[self.mro_rsrp_list[max(temp_value_list)]+48] += 1
            temp_mro_earfcn[self.mro_rsrp_list[int(temp_value[0])]] += 1

            ecid_earfcn = '_'.join(map(str, (ecid1, temp_value[11])))
            try:
                ecid_operator = '_'.join(map(str, (ecid1, self.config_mro['operator_list'][int(temp_value[11])])))
            except:
                ecid_operator = '_'.join(map(str, (ecid1, '其它')))

            try:
                self.temp_mro_data[report_time]['mro_earfcn'][ecid_earfcn] += numpy.array(temp_mro_earfcn)
            except:
                try:
                    self.temp_mro_data[report_time]['mro_earfcn'][ecid_earfcn] = numpy.array(temp_mro_earfcn)
                except:
                    try:
                        self.temp_mro_data[report_time]['mro_earfcn'] = {}
                        self.temp_mro_data[report_time]['mro_earfcn'][ecid_earfcn] = numpy.array(temp_mro_earfcn)
                    except:
                        self.temp_mro_data[report_time] = {}
                        self.temp_mro_data[report_time]['mro_earfcn'] = {}
                        self.temp_mro_data[report_time]['mro_earfcn'][ecid_earfcn] = numpy.array(temp_mro_earfcn)

            try:
                self.temp_mro_data[report_time]['mro_earfcn_operator'][ecid_operator] += numpy.array(temp_mro_earfcn)
            except:
                try:
                    self.temp_mro_data[report_time]['mro_earfcn_operator'][ecid_operator] = numpy.array(temp_mro_earfcn)
                except:
                    try:
                        self.temp_mro_data[report_time]['mro_earfcn_operator'] = {}
                        self.temp_mro_data[report_time]['mro_earfcn_operator'][ecid_operator] = numpy.array(temp_mro_earfcn)
                    except:
                        self.temp_mro_data[report_time] = {}
                        self.temp_mro_data[report_time]['mro_earfcn_operator'] = {}
                        self.temp_mro_data[report_time]['mro_earfcn_operator'][ecid_operator] = numpy.array(temp_mro_earfcn)

    def get_report_time(self, tree):

        """获取报告时段"""

        report_time = '-'
        if 'hour' in self.config_main['gather_type']:
            for temp_file_header in tree.iter(tag='fileHeader'):
                report_time = temp_file_header.attrib['startTime'][
                              :temp_file_header.attrib['startTime'].find(':')]
                break
        return report_time

    def get_enbid(self, tree):

        """获取enbid"""
        enbid = ''
        for temp_file_header in tree.iter(tag='eNB'):
            enbid = temp_file_header.attrib['id']
            break
        return enbid

    def parse_process(self, mr_type):

        """多进程控制"""

        # 获取表头
        for n in self.parse_file_list[mr_type]:
            for o in self.parse_file_list[mr_type][n]:
                self.child_parse_process(mr_type, n, o, ishead=1)
                break
            break

        # 统计个数
        self.num_files = 0
        # 统计文件个数
        for temp_n in self.parse_file_list[mr_type]:
            self.num_files += len(self.parse_file_list[mr_type][temp_n])
        # parse_file_list 格式为 {'mrs':{'gz':[...],'xml':[...]},
        #                         'mro':{'gz':[...],'xml':[...]},
        #                        }
        # temp_i 为文件类型（gz or xml）
        # temp_j 为待处理文件
        process_manager = multiprocessing.Manager()
        process_queue = process_manager.Queue()
        # process_lock = process_manager.Lock()
        process_pool = multiprocessing.Pool(processes=int(self.config_main['process'][0]))
        process_listen = process_pool.apply_async(self.listen, args=(process_queue, mr_type,))
        jobs = []
        for n in self.parse_file_list[mr_type]:
            for o in self.parse_file_list[mr_type][n]:
                job = process_pool.apply_async(self.child_parse_process, args=(mr_type, n, o, process_queue))
                jobs.append(job)
        for job in jobs:
            job.get()

        process_queue.put('all_finish')
        process_pool.close()
        process_pool.join()
        # 汇总子进程数据
        if mr_type == 'mrs':
            self.mrs_data_data = process_listen.get()
        elif mr_type == 'mro':
            self.mro_data_data = process_listen.get()

    def filter(self, file_name, type_filter):
        filter_id = self.config_filter['filter_id']
        filter_type = self.config_filter['filter_type']
        filter_hour = self.config_filter['filter_hour']
        temp_filter_id = os.path.split(file_name)[-1][19:25]
        temp_filter_type = os.path.split(file_name)[-1][7:10].lower()
        temp_filter_hour = os.path.split(file_name)[-1][34:36]

        if (temp_filter_id in filter_id or filter_id == ['']) and (
                        temp_filter_type in filter_type or filter_type == ['']) and (
                        temp_filter_hour in filter_hour or filter_hour == ['']):
            if type_filter == 'xml' or type_filter == 'gz':
                if os.path.isfile(os.path.join(self.config_main['target_path'][0], os.path.split(file_name)[-1])):
                    pass
                else:
                    shutil.copyfile(file_name, os.path.join(self.config_main['target_path'][0], os.path.split(
                        file_name)[-1]))
            return 1
        else:
            return 0

    def child_parse_process(self, mr_type, file_type, file_name, queue='', ishead=0):

        """文件格式判断、解压、parse xml"""
        log_file_child_num = 0
        log_file_child_list = []
        if file_type == 'xml':
            try:
                if self.config_filter['active_filter'] != ['1']:
                    tree = ET.parse(file_name)
                    self.parser(tree, mr_type, queue, ishead)
                    log_file_child_num += 1
                    log_file_child_list.append(file_name)
                else:
                    if self.filter(file_name, 'xml') == 1:
                        tree = ET.parse(file_name)
                        self.parser(tree, mr_type, queue, ishead)
                        log_file_child_num += 1
                        log_file_child_list.append(file_name)
            except:
                pass
                # traceback.print_exc()

        elif file_type == 'gz':
            try:
                try:
                    tar_f = tarfile.open(file_name)
                    for temp_file in tar_f.getnames():
                        try:
                            temp_file_tar_f = tar_f.extractfile(temp_file)
                            temp_file_suffix = temp_file.split('.')[-1].lower()
                            if temp_file_suffix == 'gz':
                                if self.config_filter['active_filter'] != ['1']:
                                    tree = ET.parse(gzip.open(temp_file_tar_f))
                                    self.parser(tree, mr_type, queue, ishead)
                                    log_file_child_num += 1
                                    log_file_child_list.append(temp_file)
                                else:
                                    if self.filter(temp_file, 'tar_gz') == 1:
                                        tree = ET.parse(gzip.open(temp_file_tar_f))
                                        self.parser(tree, mr_type, queue, ishead)
                                        tar_f.extract(temp_file, self.config_main['target_path'][0])
                                        log_file_child_num += 1
                                        log_file_child_list.append(temp_file)

                            elif temp_file_suffix == 'xml':
                                if self.config_filter['active_filter'] != ['1']:
                                    tree = ET.parse(temp_file_tar_f)
                                    self.parser(tree, mr_type, queue, ishead)
                                    log_file_child_num += 1
                                    log_file_child_list.append(temp_file)
                                else:
                                    if self.filter(temp_file, 'tar_gz') == 1:
                                        tree = ET.parse(temp_file_tar_f)
                                        self.parser(tree, mr_type, queue, ishead)
                                        tar_f.extract(temp_file, self.config_main['target_path'][0])
                                        log_file_child_num += 1
                                        log_file_child_list.append(temp_file)
                        except:
                            pass
                            # traceback.print_exc()

                    tar_f.close()
                except:
                    # traceback.print_exc()
                    gzip_file = gzip.open(file_name)
                    if self.config_filter['active_filter'] != ['1']:
                        tree = ET.parse(gzip_file)
                        self.parser(tree, mr_type, queue, ishead)
                        log_file_child_num += 1
                        log_file_child_list.append(file_name)
                    else:
                        if self.filter(file_name, 'gz') == 1:
                            tree = ET.parse(gzip_file)
                            self.parser(tree, mr_type, queue, ishead)
                            log_file_child_num += 1
                            log_file_child_list.append(file_name)
            except:
                traceback.print_exc()

        # 数据送到queue
        if ishead == 0:
            type_list = {'mrs': self.temp_mrs_data,
                         'mro': self.temp_mro_data}
            self.queue_send(queue, type_list[mr_type])
            # 发送进度条及文件名称，为log；
            # queue.put('prog')
            queue.put(['prog', mr_type, file_name, log_file_child_num, log_file_child_list])
            type_list[mr_type] = {}

    def parser(self, tree, mr_type, queue, ishead):
        if ishead == 0:
            if mr_type == 'mrs':
                report_time = self.get_report_time(tree)
                for temp_mr_name in tree.iter('measurement'):
                    temp_table_name_1 = temp_mr_name.attrib['mrName']
                    if temp_table_name_1 in self.mrs_parse_sheet:
                        for temp_id in temp_mr_name.iter('object'):
                            temp_mrs_ecid = temp_id.attrib['id']
                            for temp_value in temp_id.iter('v'):
                                temp_values = numpy.array(list(map(float, temp_value.text.rstrip().split(' '))))
                                try:
                                    self.temp_mrs_data[report_time][temp_table_name_1][temp_mrs_ecid] += temp_values
                                except:
                                    try:
                                        self.temp_mrs_data[report_time][temp_table_name_1][temp_mrs_ecid] = temp_values
                                    except:
                                        try:
                                            self.temp_mrs_data[report_time][temp_table_name_1] = {}
                                            self.temp_mrs_data[report_time][temp_table_name_1][
                                                temp_mrs_ecid] = temp_values
                                        except:
                                            self.temp_mrs_data[report_time] = {}
                                            self.temp_mrs_data[report_time][temp_table_name_1] = {}
                                            self.temp_mrs_data[report_time][temp_table_name_1][
                                                temp_mrs_ecid] = temp_values

            elif mr_type == 'mro':
                table_list = {'mro_main': self.mro_main,
                              'mro_ecid': self.mro_ecid,
                              'mro_rsrp': self.mro_rsrp,
                              'mro_rsrp_mdt': self.mro_rsrp_mdt,
                              'mro_aoa': self.mro_aoa,
                              'mro_earfcn': self.mro_earfcn,
                              'mro_report_num': self.mro_report,
                              }
                report_time = self.get_report_time(tree)
                enbid = self.get_enbid(tree)
                for measurement in tree.iter('measurement'):
                    for smr in measurement.iter('smr'):
                        head_temp = smr.text.replace('.', '_').rstrip().split(' ')
                        if head_temp[0] == 'MR_LteScRSRP':
                            for object_mro in measurement.iter('object'):
                                # 分别生成需要的mro表
                                for table_temp in self.config_mro['mro_parse_sheet']:
                                    if table_temp != 'mro_ecid':
                                        table_list[table_temp](object_mro, report_time, enbid)
                                    else:
                                        table_list[table_temp](object_mro, '-', enbid)

        elif ishead == 1:
            # 获取需处理表名
            if mr_type == 'mrs':
                for mrname in tree.iter('measurement'):
                    if (self.config_mrs[
                            'mrs_parse_sheet'
                        ] == [''] or mrname.attrib['mrName'] in self.config_mrs['mrs_parse_sheet']) and mrname.attrib[
                        'mrName'
                    ] not in self.config_mrs['mrs_exception_sheet']:
                        self.mrs_parse_sheet.append(mrname.attrib['mrName'])

            # 获取mrs表头
            if mr_type == 'mrs':
                for mrname in tree.iter('measurement'):
                    temp_table_name = mrname.attrib['mrName']
                    if temp_table_name in self.mrs_parse_sheet:
                        for smr in mrname.iter('smr'):
                            head_temp = smr.text.replace('.', '_').rstrip().split(' ')
                            self.mrs_head[temp_table_name] = head_temp

    def queue_send(self, queue, data):
        for temp_report_time in data:
            for temp_table in data[temp_report_time]:
                for temp_ecid in data[temp_report_time][temp_table]:
                    queue.put(['data', [temp_report_time,
                                        temp_table,
                                        temp_ecid,
                                        data[temp_report_time][temp_table][temp_ecid]
                                        ]
                               ]
                              )

    def gather(self, mr_type):
        gather_data = {'mrs': {},
                       'mro': {}
                       }
        temp_gather_data = {}
        if mr_type == 'mrs':
            temp_gather_data = self.mrs_data_data
        elif mr_type == 'mro':
            temp_gather_data = self.mro_data_data
        for temp_table in temp_gather_data[mr_type]:
            # mro_ecid 表不生成小时级，汇总阶段直接删除此表数据
            if temp_table == 'mro_ecid':
                temp_gather_data[mr_type][temp_table] = []
            else:
                if temp_table not in gather_data[mr_type]:
                    gather_data[mr_type][temp_table] = {'-': {}}
                for temp_report_time in temp_gather_data[mr_type][temp_table]:
                    for temp_ecid in temp_gather_data[mr_type][temp_table][temp_report_time]:
                        try:
                            gather_data[mr_type][temp_table]['-'][temp_ecid] += temp_gather_data[
                                mr_type][temp_table][temp_report_time][temp_ecid]
                        except:
                            gather_data[mr_type][temp_table]['-'][temp_ecid] = temp_gather_data[
                                mr_type][temp_table][temp_report_time][temp_ecid]
        if mr_type == 'mrs':
            self.mrs_data_data['mrs'] = gather_data['mrs']
        elif mr_type == 'mro':
            self.mro_data_data['mro'] = gather_data['mro']

    def listen(self, queue, mr_type):
        # 表结构
        # 表名:{time:
        #               {ecid:[值1，值2，......]
        #                   }
        #       }
        all_list = {'mrs': {},
                    'mro': {}
                    }
        num_ii = 0
        self.progress(self.all_num[mr_type], num_ii)
        # 生成log文件
        f_log_csv = open(''.join((self.config_main['target_path'][0],
                                  '/LOG_Parse_File_List.csv'
                                  )), 'a', encoding='utf-8-sig'
                         )
        while 1:
            value = queue.get()
            if value[0] == 'data':
                report_time = value[1][0]
                table_name = value[1][1]
                table_id = value[1][2]
                table_value = numpy.array(value[1][3])
                try:
                    all_list[mr_type][table_name][report_time][table_id] += table_value
                except:
                    try:
                        all_list[mr_type][table_name][report_time][table_id] = table_value
                    except:
                        try:
                            all_list[mr_type][table_name][report_time] = {}
                            all_list[mr_type][table_name][report_time][table_id] = table_value
                        except:
                            all_list[mr_type][table_name] = {}
                            all_list[mr_type][table_name][report_time] = {}
                            all_list[mr_type][table_name][report_time][table_id] = table_value
            elif value[0] == 'prog':
                num_ii += 1
                self.progress(self.all_num[mr_type], num_ii)
                # 记录log
                for temp_file in value[4]:
                    f_log_csv.write(value[1])
                    f_log_csv.write(',')
                    f_log_csv.write(os.path.split(value[2])[-1])
                    f_log_csv.write(',')
                    f_log_csv.write(str(value[3]))
                    f_log_csv.write(',')
                    f_log_csv.write(os.path.split(temp_file)[-1])
                    f_log_csv.write('\n')
            elif value == 'all_finish':
                f_log_csv.close()
                return all_list

    def writer(self, mr_type, time_type):
        if self.config_main['timing'][0] == '1':
            temp_day_head = ''.join(('_', self.yesterday))
            temp_day = self.yesterday
        else:
            temp_day_head = ''
            temp_day = '-'
        if mr_type == 'mrs':
            for table_mrs in self.mrs_data_data['mrs']:
                with open(os.path.join(self.config_main['target_path'][0],
                                       '{0}{1}_{2}.csv'.format(table_mrs, temp_day_head, time_type)), 'w',
                          newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    if table_mrs == 'MR.RSRP':
                        writer.writerow(['DAY', 'TIME', 'ECID', 'ENB_ID', 'ENB_CELLID', 'MR覆盖率（RSRP>=-110)',
                                         'RSRP>=-110计数器', 'ALL计数器'] + self.mrs_head[table_mrs])
                        # print(self.mrs_data_data)
                        for temp_report_time in self.mrs_data_data['mrs'][table_mrs]:
                            for temp_ecid in self.mrs_data_data['mrs'][table_mrs][temp_report_time]:
                                enb_cellid = '_'.join((str(int(temp_ecid) // 256), str(int(temp_ecid) % 256)))
                                numerator = numpy.sum(
                                    self.mrs_data_data['mrs'][table_mrs][temp_report_time][temp_ecid][7:]
                                )
                                denominator = numpy.sum(
                                    self.mrs_data_data['mrs'][table_mrs][temp_report_time][temp_ecid]
                                )
                                if denominator == 0:
                                    range_mrs = '-'
                                else:
                                    range_mrs = round(numerator / denominator * 100, 2)
                                writer.writerow([temp_day, temp_report_time, temp_ecid, int(temp_ecid) // 256,
                                                 enb_cellid, range_mrs, numerator, denominator] + list(
                                    self.mrs_data_data['mrs'][table_mrs][temp_report_time][temp_ecid]))
                    else:
                        writer.writerow(['DAY', 'TIME', 'ECID', ] + self.mrs_head[table_mrs])
                        # print(self.mrs_data_data)
                        for temp_report_time in self.mrs_data_data['mrs'][table_mrs]:
                            for temp_ecid in self.mrs_data_data['mrs'][table_mrs][temp_report_time]:
                                writer.writerow([temp_day,
                                                 temp_report_time,
                                                 temp_ecid,
                                                 ] + list(
                                    self.mrs_data_data['mrs'][table_mrs][temp_report_time][temp_ecid]))
        elif mr_type == 'mro':

            for table in self.mro_data_data['mro']:
                if table == 'mro_main':
                    with open(os.path.join(self.config_main['target_path'][0],
                                           '{0}{1}_{2}.csv'.format(table,
                                                                   temp_day_head,
                                                                   time_type)), 'w', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(('DAY', 'TIME', 'ECID', 'ENBID', 'ENB_CLEEID', 'EARFCN', 'PCI', 'MR.LteScRSRP',
                                         'MR.LteScRSRQ', '覆盖距离（m）', 'MR.LteScPHR', 'MR.LteScAOA',
                                         'MR.LteScSinrUL',
                                         'MR.LteScPUSCHPRBNum', 'MR.LteScPDSCHPRBNum', 's_samplint',
                                         'RSRP>=-110采样点', 'MR覆盖率(RSRP>=-110)',
                                         'CMCC重叠覆盖采样点', 'CMCC同频重叠覆盖率'))
                        for temp_report_time in self.mro_data_data['mro'][table]:
                            for table_id in self.mro_data_data['mro'][table][temp_report_time]:
                                temp_value = self.mro_data_data['mro'][table][temp_report_time][table_id]
                                if temp_value[8] != 0:
                                    temp_value_1 = temp_value[0:8] / temp_value[8]
                                    temp_value_1[0] = temp_value_1[0] - 140
                                    temp_value_1[2] = temp_value_1[2] * 78
                                else:
                                    temp_value_1 = temp_value[0:8]
                                temp_value_total = list(temp_value_1) + list(temp_value[8:9])
                                temp_value_total = list(map(int, temp_value_total))
                                temp_value = temp_value_total + list(
                                    [temp_value[9],
                                     round(temp_value[9] / temp_value[8]*100, 2)]
                                ) + list(
                                    [temp_value[10],
                                     round(temp_value[10]/temp_value[8]*100, 2)]
                                )
                                temp_id = table_id.split('_')
                                temp_senbid = str(int(temp_id[0]) // 256)
                                writer.writerow([temp_day,
                                                 temp_report_time,
                                                 temp_id[0],
                                                 temp_senbid,
                                                 '_'.join((temp_senbid, str(int(temp_id[0]) % 256))),
                                                 temp_id[1],
                                                 temp_id[2]
                                                 ] + temp_value
                                                )
                elif table == 'mro_ecid':
                    with open(os.path.join(self.config_main['target_path'][0],
                                           '{0}{1}_{2}.csv'.format('mro_ecid',
                                                                   temp_day_head,
                                                                   'sum')), 'w', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(('DAY', 'TIME', 'S_ECID', 'ENBID', 'ENB_CELLID', 'S_EARFCN', 'S_PCI',
                                         'N_ENB_CELLID', 'N_EARFCN', 'N_PCI', 'Distance(m)', 'Scrsrp', 'ScTadv',
                                         'Ncrsrp', ' N_Samplint', '<-10db', '-10db', '-9db', '-8db', '-7db',
                                         '-6db', '-5db', '-4db', '-3db', '-2db', '-1db', '0db', '1db', '2db', '3db',
                                         '4db', '5db', ' 6db', '7db', '8db', '9db', '10db', '>10db'))
                        for temp_report_time in self.mro_data_data['mro'][table]:
                            for table_id in self.mro_data_data['mro'][table][temp_report_time]:
                                temp_value = self.mro_data_data['mro'][table][temp_report_time][table_id]
                                if temp_value[3] != 0:
                                    temp_value_1 = temp_value[0:3] / temp_value[3]
                                else:
                                    temp_value_1 = temp_value[0:3]
                                temp_value = list(temp_value_1) + list(temp_value[3:])
                                temp_value = list(map(float, temp_value))
                                temp_id = table_id.split('_')
                                temp_senbid = str(int(temp_id[0]) // 256)
                                temp_ncellid, temp_min_stance = self.min_distance_cell(
                                    temp_senbid, temp_id[3], temp_id[4]
                                )
                                writer.writerow([temp_day, temp_report_time, temp_id[0], temp_senbid,
                                                 '_'.join((str(int(temp_id[0]) // 256), str(int(temp_id[0]) % 256))),
                                                 temp_id[1], temp_id[2], temp_ncellid, temp_id[3], temp_id[4],
                                                 temp_min_stance] + temp_value)

                    with open(os.path.join(self.config_main['target_path'][0],
                                           '{0}{1}_{2}.csv'.format('mro_ecid_yuan',
                                                                   temp_day_head,
                                                                   'sum')), 'w', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow((
                            'ECI_ECI', 'S_ENBID', 'S_CELLID', 'S_PCI', 'S_EARFCN', 'A_ENBID', 'A_CELLID', 'A_PCI',
                            'A_EARFCN', 'distance', 'total', 'GE-10db', 'GE-6db', 'NbrAvg', 'NbrMax', 'NbrMin'
                        ))
                        for temp_report_time in self.mro_data_data['mro'][table]:
                            for table_id in self.mro_data_data['mro'][table][temp_report_time]:
                                temp_id = table_id.split('_')
                                temp_senbid = str(int(temp_id[0]) // 256)
                                temp_scellid = str(int(temp_id[0]) % 256)
                                temp_ncellid, temp_min_stance = self.min_distance_cell(
                                    temp_senbid, temp_id[3], temp_id[4]
                                )
                                if temp_ncellid != '-':
                                    temp_N_ECI = temp_ncellid.split('_')
                                    N_ECI = str(int(temp_N_ECI[0])*256+int(temp_N_ECI[1]))
                                    N_enbid = temp_N_ECI[0]
                                    N_cellid = temp_N_ECI[1]
                                else:
                                    continue
                                    # N_ECI = '-'
                                    # N_enbid = '-'
                                    # N_cellid = '-'
                                temp_ECI_ECI = '_'.join((temp_id[0], N_ECI))
                                temp_value = self.mro_data_data['mro'][table][temp_report_time][table_id]
                                if temp_value[3] != 0:
                                    temp_value_1 = temp_value[2] // temp_value[3]
                                else:
                                    temp_value_1 = temp_value[2]
                                writer.writerow([temp_ECI_ECI,
                                                 temp_senbid,
                                                 temp_scellid,
                                                 temp_id[2],
                                                 temp_id[1],
                                                 N_enbid,
                                                 N_cellid,
                                                 temp_id[4],
                                                 temp_id[3],
                                                 temp_min_stance,
                                                 temp_value[3],
                                                 sum(temp_value[5:26]),
                                                 sum(temp_value[9:22]),
                                                 temp_value_1 - 140
                                                 ])
                elif table in ('mro_rsrp', 'mro_rsrp_mdt'):
                    with open(os.path.join(self.config_main['target_path'][0],
                                           '{0}{1}_{2}.csv'.format(table,
                                                                   temp_day_head,
                                                                   time_type)), 'w', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        if table == 'mro_rsrp':
                            writer.writerow(('DAY', 'TIME', 'ECID', 'ENBID', 'ENB_CELLID', 'MR覆盖率（RSRP>=-110)',
                                             'RSRP>=-110计数器', 'ALL计数器',
                                             'MR_RSRP_00', 'MR_RSRP_01', 'MR_RSRP_02', 'MR_RSRP_03', 'MR_RSRP_04',
                                             'MR_RSRP_05', 'MR_RSRP_06', 'MR_RSRP_07', 'MR_RSRP_08', 'MR_RSRP_09',
                                             'MR_RSRP_10', 'MR_RSRP_11', 'MR_RSRP_12', 'MR_RSRP_13', 'MR_RSRP_14',
                                             'MR_RSRP_15', 'MR_RSRP_16', 'MR_RSRP_17', 'MR_RSRP_18', 'MR_RSRP_19',
                                             'MR_RSRP_20', 'MR_RSRP_21', 'MR_RSRP_22', 'MR_RSRP_23', 'MR_RSRP_24',
                                             'MR_RSRP_25', 'MR_RSRP_26', 'MR_RSRP_27', 'MR_RSRP_28', 'MR_RSRP_29',
                                             'MR_RSRP_30', 'MR_RSRP_31', 'MR_RSRP_32', 'MR_RSRP_33', 'MR_RSRP_34',
                                             'MR_RSRP_35', 'MR_RSRP_36', 'MR_RSRP_37', 'MR_RSRP_38', 'MR_RSRP_39',
                                             'MR_RSRP_40', 'MR_RSRP_41', 'MR_RSRP_42', 'MR_RSRP_43', 'MR_RSRP_44',
                                             'MR_RSRP_45', 'MR_RSRP_46', 'MR_RSRP_47'))
                        elif table == 'mro_rsrp_mdt':
                            writer.writerow(('DAY', 'TIME', 'ECID', 'ENBID', 'ENB_CELLID', 'MR覆盖率（RSRP>=-110)',
                                             'RSRP>=-110计数器', 'ALL计数器',
                                             'MR_RSRP_00', 'MR_RSRP_01', 'MR_RSRP_02', 'MR_RSRP_03', 'MR_RSRP_04',
                                             'MR_RSRP_05', 'MR_RSRP_06', 'MR_RSRP_07', 'MR_RSRP_08', 'MR_RSRP_09',
                                             'MR_RSRP_10', 'MR_RSRP_11', 'MR_RSRP_12', 'MR_RSRP_13', 'MR_RSRP_14',
                                             'MR_RSRP_15', 'MR_RSRP_16', 'MR_RSRP_17', 'MR_RSRP_18', 'MR_RSRP_19',
                                             'MR_RSRP_20', 'MR_RSRP_21', 'MR_RSRP_22', 'MR_RSRP_23', 'MR_RSRP_24',
                                             'MR_RSRP_25', 'MR_RSRP_26', 'MR_RSRP_27', 'MR_RSRP_28', 'MR_RSRP_29',
                                             'MR_RSRP_30', 'MR_RSRP_31', 'MR_RSRP_32', 'MR_RSRP_33', 'MR_RSRP_34',
                                             'MR_RSRP_35', 'MR_RSRP_36', 'MR_RSRP_37', 'MR_RSRP_38', 'MR_RSRP_39',
                                             'MR_RSRP_40', 'MR_RSRP_41', 'MR_RSRP_42', 'MR_RSRP_43', 'MR_RSRP_44',
                                             'MR_RSRP_45', 'MR_RSRP_46', 'MR_RSRP_47',
                                             'ue方位角偏差超过门限计数器', 'ue方位角偏差超过门限占比', 'ue平均方位角'))

                        for temp_report_time in self.mro_data_data['mro'][table]:
                            for ecid_id in self.mro_data_data['mro'][table][temp_report_time]:
                                temp_value = self.mro_data_data['mro'][table][temp_report_time][ecid_id]
                                temp_value_1 = sum(temp_value[7:47])
                                temp_value_2 = sum(temp_value[0:47])
                                if temp_value_2 != 0:
                                    temp_value_3 = round(temp_value_1 / temp_value_2 * 100, 2)
                                else:
                                    temp_value_3 = '-'

                                temp_enbid = str(int(ecid_id) // 256)
                                temp_enb_cellid = '_'.join((str(int(ecid_id) // 256), str(int(ecid_id) % 256)))
                                if table == 'mro_rsrp':
                                    writer.writerow([temp_day, temp_report_time, ecid_id, temp_enbid, temp_enb_cellid,
                                                     temp_value_3,
                                                     temp_value_1,
                                                     temp_value_2] + list(temp_value))
                                elif table == 'mro_rsrp_mdt':
                                    writer.writerow(
                                        [temp_day, temp_report_time, ecid_id, temp_enbid, temp_enb_cellid,
                                                     temp_value_3,
                                                     temp_value_1,
                                                     temp_value_2] + list(temp_value[:-1]) + [
                                            round(temp_value[48] / temp_value_2 * 100, 2),
                                            temp_value[49] // temp_value_2
                                        ]
                                    )
                elif table == 'mro_rsrp_mdt_details':
                    with open(os.path.join(self.config_main['target_path'][0],
                                           '{0}{1}_{2}.csv'.format(table,
                                                                   temp_day_head,
                                                                   time_type)), 'w', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(('DAY', 'TIME', 'ECID', 'ENBID', 'ENB_CELLID', 'CELL_NAME', 'lon',
                                         'lat', 'SC_RSRP', 'LteScRSRQ', 'LteScSinrUL', 'LteScTadv', 'SAMPLES',
                                         'ue所在位置超出小区方向', 'ue方向角'))
                        for temp_report_time in self.mro_data_data['mro'][table]:
                            for ecid_id in self.mro_data_data['mro'][table][temp_report_time]:
                                temp_value = ecid_id.split('_')
                                temp_value_num = list(self.mro_data_data['mro'][table][temp_report_time][ecid_id])
                                temp_enbid = str(int(temp_value[0]) // 256)
                                temp_enb_cellid = '_'.join((str(int(temp_value[0]) // 256),
                                                            str(int(temp_value[0]) % 256)))
                                try:
                                    temp_enb_cellid_cn_name = self.mro_enbcellid_cn_name[temp_enb_cellid]
                                except:
                                    temp_enb_cellid_cn_name = ''
                                try:
                                    temp_value_num[2] = round(float(temp_value_num[2])/float(temp_value_num[1]),2)
                                except:
                                    pass
                                writer.writerow([temp_day, temp_report_time, temp_value[0],
                                                 temp_enbid,
                                                 temp_enb_cellid,
                                                 temp_enb_cellid_cn_name] + temp_value[1:] + temp_value_num)
                elif table == 'mro_mdt_overlap':
                    with open(os.path.join(self.config_main['target_path'][0],
                                           '{0}{1}_{2}.csv'.format(table,
                                                                   temp_day_head,
                                                                   time_type)), 'w', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(('Report_Time', 'UELongitude', 'UELatitude', 'EARFCN', 'ECID', 'S_RSRP',
                                         '重叠覆盖邻区数','点出现次数'))
                        for temp_report_time in self.mro_data_data['mro'][table]:
                            for ecid_id in self.mro_data_data['mro'][table][temp_report_time]:
                                temp_value = ecid_id.split('_')
                                writer.writerow([temp_report_time,]+temp_value + [self.mro_data_data['mro'][table][
                                                                                    temp_report_time][
                                                                   ecid_id],])
                elif table == 'mro_aoa':
                    with open(os.path.join(self.config_main['target_path'][0],
                                           '{0}{1}_{2}.csv'.format(table,
                                                                   temp_day_head,
                                                                   time_type)), 'w', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(('DAY', 'TIME', 'ECID', 'ENBID', 'ENB_CELLID',
                                         'MR_AOA_00', 'MR_AOA_01', 'MR_AOA_02', 'MR_AOA_03', 'MR_AOA_04', 'MR_AOA_05',
                                         'MR_AOA_06', 'MR_AOA_07', 'MR_AOA_08', 'MR_AOA_09', 'MR_AOA_10', 'MR_AOA_11',
                                         'MR_AOA_12', 'MR_AOA_13', 'MR_AOA_14', 'MR_AOA_15', 'MR_AOA_16', 'MR_AOA_17',
                                         'MR_AOA_18', 'MR_AOA_19', 'MR_AOA_20', 'MR_AOA_21', 'MR_AOA_22', 'MR_AOA_23',
                                         'MR_AOA_24', 'MR_AOA_25', 'MR_AOA_26', 'MR_AOA_27', 'MR_AOA_28', 'MR_AOA_29',
                                         'MR_AOA_30', 'MR_AOA_31', 'MR_AOA_32', 'MR_AOA_33', 'MR_AOA_34', 'MR_AOA_35',
                                         'MR_AOA_36', 'MR_AOA_37', 'MR_AOA_38', 'MR_AOA_39', 'MR_AOA_40', 'MR_AOA_41',
                                         'MR_AOA_42', 'MR_AOA_43', 'MR_AOA_44', 'MR_AOA_45', 'MR_AOA_46', 'MR_AOA_47',
                                         'MR_AOA_48', 'MR_AOA_49', 'MR_AOA_50', 'MR_AOA_51', 'MR_AOA_52', 'MR_AOA_53',
                                         'MR_AOA_54', 'MR_AOA_55', 'MR_AOA_56', 'MR_AOA_57', 'MR_AOA_58', 'MR_AOA_59',
                                         'MR_AOA_60', 'MR_AOA_61', 'MR_AOA_62', 'MR_AOA_63', 'MR_AOA_64', 'MR_AOA_65',
                                         'MR_AOA_66', 'MR_AOA_67', 'MR_AOA_68', 'MR_AOA_69', 'MR_AOA_70', 'MR_AOA_71'))
                        for temp_report_time in self.mro_data_data['mro'][table]:
                            for ecid_id in self.mro_data_data['mro'][table][temp_report_time]:
                                temp_value = self.mro_data_data['mro'][table][temp_report_time][ecid_id]
                                temp_enbid = str(int(ecid_id) // 256)
                                temp_enb_cellid = '_'.join((str(int(ecid_id) // 256), str(int(ecid_id) % 256)))
                                writer.writerow([temp_day, temp_report_time, ecid_id, temp_enbid, temp_enb_cellid
                                                 ] + list(temp_value))
                elif table == 'mro_earfcn':
                    with open(os.path.join(self.config_main['target_path'][0],
                                           '{0}{1}_{2}.csv'.format(table,
                                                                   temp_day_head,
                                                                   time_type)), 'w', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(
                            ['DAY', 'TIME', 'ECID', 'ENBID', 'ENB_CELLID', 'n_EARFCN',
                             'S_MR覆盖率（RSRP>=-110)', 'S_RSRP>=-110计数器', 'S_ALL计数器',
                             'N_MR覆盖率（RSRP>=-110)', 'N_RSRP>=-110计数器', 'N_ALL计数器',
                             's_RSRP_00', 's_RSRP_01',
                             's_RSRP_02', 's_RSRP_03', 's_RSRP_04', 's_RSRP_05', 's_RSRP_06', 's_RSRP_07', 's_RSRP_08',
                             's_RSRP_09', 's_RSRP_10', 's_RSRP_11', 's_RSRP_12', 's_RSRP_13', 's_RSRP_14', 's_RSRP_15',
                             's_RSRP_16', 's_RSRP_17', 's_RSRP_18', 's_RSRP_19', 's_RSRP_20', 's_RSRP_21', 's_RSRP_22',
                             's_RSRP_23', 's_RSRP_24', 's_RSRP_25', 's_RSRP_26', 's_RSRP_27', 's_RSRP_28', 's_RSRP_29',
                             's_RSRP_30', 's_RSRP_31', 's_RSRP_32', 's_RSRP_33', 's_RSRP_34', 's_RSRP_35', 's_RSRP_36',
                             's_RSRP_37', 's_RSRP_38', 's_RSRP_39', 's_RSRP_40', 's_RSRP_41', 's_RSRP_42', 's_RSRP_43',
                             's_RSRP_44', 's_RSRP_45', 's_RSRP_46', 's_RSRP_47',
                             'n_RSRP_00', 'n_RSRP_01', 'n_RSRP_02',
                             'n_RSRP_03', 'n_RSRP_04', 'n_RSRP_05', 'n_RSRP_06', 'n_RSRP_07', 'n_RSRP_08', 'n_RSRP_09',
                             'n_RSRP_10', 'n_RSRP_11', 'n_RSRP_12', 'n_RSRP_13', 'n_RSRP_14', 'n_RSRP_15', 'n_RSRP_16',
                             'n_RSRP_17', 'n_RSRP_18', 'n_RSRP_19', 'n_RSRP_20', 'n_RSRP_21', 'n_RSRP_22', 'n_RSRP_23',
                             'n_RSRP_24', 'n_RSRP_25', 'n_RSRP_26', 'n_RSRP_27', 'n_RSRP_28', 'n_RSRP_29', 'n_RSRP_30',
                             'n_RSRP_31', 'n_RSRP_32', 'n_RSRP_33', 'n_RSRP_34', 'n_RSRP_35', 'n_RSRP_36', 'n_RSRP_37',
                             'n_RSRP_38', 'n_RSRP_39', 'n_RSRP_40', 'n_RSRP_41', 'n_RSRP_42', 'n_RSRP_43', 'n_RSRP_44',
                             'n_RSRP_45', 'n_RSRP_46', 'n_RSRP_47'])
                        for temp_report_time in self.mro_data_data['mro'][table]:
                            for ecid_id in self.mro_data_data['mro'][table][temp_report_time]:
                                temp_value = self.mro_data_data['mro'][table][temp_report_time][ecid_id]
                                s_temp_value_1 = sum(temp_value[7:48])
                                s_temp_value_2 = sum(temp_value[0:48])
                                if s_temp_value_2 != 0:
                                    s_temp_value_3 = round(s_temp_value_1 / s_temp_value_2 * 100, 2)
                                else:
                                    s_temp_value_3 = '-'
                                n_temp_value_1 = sum(temp_value[55:])
                                n_temp_value_2 = sum(temp_value[48:])
                                if n_temp_value_2 != 0:
                                    n_temp_value_3 = round(n_temp_value_1 / n_temp_value_2 * 100, 2)
                                else:
                                    n_temp_value_3 = '-'
                                temp_ecid_earfcn = ecid_id.split('_')
                                temp_enbid = str(int(temp_ecid_earfcn[0]) // 256)
                                temp_enb_cellid = '_'.join((str(int(temp_ecid_earfcn[0]) // 256), str(int(temp_ecid_earfcn[0]) % 256)))
                                writer.writerow([temp_day, temp_report_time, temp_ecid_earfcn[0], temp_enbid,
                                                 temp_enb_cellid,
                                                 temp_ecid_earfcn[1],
                                                 s_temp_value_3, s_temp_value_1, s_temp_value_2,
                                                 n_temp_value_3, n_temp_value_1, n_temp_value_2,
                                                 ] + list(temp_value))

                elif table == 'mro_earfcn_operator':
                    with open(os.path.join(self.config_main['target_path'][0],
                                           '{0}{1}_{2}.csv'.format(table,
                                                                   temp_day_head,
                                                                   time_type)), 'w', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(
                            ['DAY', 'TIME', 'ECID', 'ENBID', 'ENB_CELLID', '运营商',
                             'S_MR覆盖率（RSRP>=-110)', 'S_RSRP>=-110计数器', 'S_ALL计数器',
                             'N_MR覆盖率（RSRP>=-110)', 'N_RSRP>=-110计数器', 'N_ALL计数器',
                             's_RSRP_00', 's_RSRP_01',
                             's_RSRP_02', 's_RSRP_03', 's_RSRP_04', 's_RSRP_05', 's_RSRP_06', 's_RSRP_07', 's_RSRP_08',
                             's_RSRP_09', 's_RSRP_10', 's_RSRP_11', 's_RSRP_12', 's_RSRP_13', 's_RSRP_14', 's_RSRP_15',
                             's_RSRP_16', 's_RSRP_17', 's_RSRP_18', 's_RSRP_19', 's_RSRP_20', 's_RSRP_21', 's_RSRP_22',
                             's_RSRP_23', 's_RSRP_24', 's_RSRP_25', 's_RSRP_26', 's_RSRP_27', 's_RSRP_28', 's_RSRP_29',
                             's_RSRP_30', 's_RSRP_31', 's_RSRP_32', 's_RSRP_33', 's_RSRP_34', 's_RSRP_35', 's_RSRP_36',
                             's_RSRP_37', 's_RSRP_38', 's_RSRP_39', 's_RSRP_40', 's_RSRP_41', 's_RSRP_42', 's_RSRP_43',
                             's_RSRP_44', 's_RSRP_45', 's_RSRP_46', 's_RSRP_47',
                             'n_RSRP_00', 'n_RSRP_01', 'n_RSRP_02',
                             'n_RSRP_03', 'n_RSRP_04', 'n_RSRP_05', 'n_RSRP_06', 'n_RSRP_07', 'n_RSRP_08', 'n_RSRP_09',
                             'n_RSRP_10', 'n_RSRP_11', 'n_RSRP_12', 'n_RSRP_13', 'n_RSRP_14', 'n_RSRP_15', 'n_RSRP_16',
                             'n_RSRP_17', 'n_RSRP_18', 'n_RSRP_19', 'n_RSRP_20', 'n_RSRP_21', 'n_RSRP_22', 'n_RSRP_23',
                             'n_RSRP_24', 'n_RSRP_25', 'n_RSRP_26', 'n_RSRP_27', 'n_RSRP_28', 'n_RSRP_29', 'n_RSRP_30',
                             'n_RSRP_31', 'n_RSRP_32', 'n_RSRP_33', 'n_RSRP_34', 'n_RSRP_35', 'n_RSRP_36', 'n_RSRP_37',
                             'n_RSRP_38', 'n_RSRP_39', 'n_RSRP_40', 'n_RSRP_41', 'n_RSRP_42', 'n_RSRP_43', 'n_RSRP_44',
                             'n_RSRP_45', 'n_RSRP_46', 'n_RSRP_47'])
                        for temp_report_time in self.mro_data_data['mro'][table]:
                            for ecid_id in self.mro_data_data['mro'][table][temp_report_time]:
                                temp_value = self.mro_data_data['mro'][table][temp_report_time][ecid_id]
                                s_temp_value_1 = sum(temp_value[7:48])
                                s_temp_value_2 = sum(temp_value[0:48])
                                if s_temp_value_2 != 0:
                                    s_temp_value_3 = round(s_temp_value_1 / s_temp_value_2 * 100, 2)
                                else:
                                    s_temp_value_3 = '-'
                                n_temp_value_1 = sum(temp_value[55:])
                                n_temp_value_2 = sum(temp_value[48:])
                                if n_temp_value_2 != 0:
                                    n_temp_value_3 = round(n_temp_value_1 / n_temp_value_2 * 100, 2)
                                else:
                                    n_temp_value_3 = '-'
                                temp_ecid_earfcn = ecid_id.split('_')
                                temp_enbid = str(int(temp_ecid_earfcn[0]) // 256)
                                temp_enb_cellid = '_'.join((str(int(temp_ecid_earfcn[0]) // 256), str(int(temp_ecid_earfcn[0]) % 256)))
                                writer.writerow([temp_day, temp_report_time, temp_ecid_earfcn[0], temp_enbid,
                                                 temp_enb_cellid,
                                                 temp_ecid_earfcn[1],
                                                 s_temp_value_3, s_temp_value_1, s_temp_value_2,
                                                 n_temp_value_3, n_temp_value_1, n_temp_value_2,
                                                 ] + list(temp_value))
                elif table == 'mro_report_num':
                    with open(os.path.join(self.config_main['target_path'][0],
                                           '{0}{1}_{2}.csv'.format(table,
                                                                   temp_day_head,
                                                                   time_type)), 'w', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(('DAY', 'TIME', 'ECID', 'ENBID', 'ENB_CELLID',
                                         '0_adj_report_num', '1_adj_report_num', '2_adj_report_num',
                                         '3_adj_report_num', 'more_than_3_adj_report_num'))
                        for temp_report_time in self.mro_data_data['mro'][table]:
                            for ecid_id in self.mro_data_data['mro'][table][temp_report_time]:
                                temp_value = self.mro_data_data['mro'][table][temp_report_time][ecid_id]
                                temp_enbid = str(int(ecid_id) // 256)
                                temp_enb_cellid = '_'.join((str(int(ecid_id) // 256), str(int(ecid_id) % 256)))
                                writer.writerow([temp_day, temp_report_time, ecid_id, temp_enbid, temp_enb_cellid
                                                 ] + list(temp_value))

    def run_manager(self, mr_type):
        logging.info(str(''.join(('>>> 解码 ', mr_type.upper(), ' 数据...'))))
        self.get_config(mr_type)
        self.parse_process(mr_type)
        logging.info(u'\n>>> {0} 计算及保存...'.format(mr_type))
        if 'hour' in self.config_main['gather_type']:
            self.writer(mr_type, 'hour')
            if 'sum' in self.config_main['gather_type']:
                self.gather(mr_type)
                self.writer(mr_type, 'sum')
        elif 'sum' in self.config_main['gather_type']:
            self.writer(mr_type, 'sum')
        logging.info('>>> {0} 数据处理完毕！'.format(mr_type))
        logging.info('-' * 26)
        logging.info('完成！{0}解码结果保存在此文件夹: {1}'.format(mr_type, self.config_main['target_path'][0]))
        logging.info('-' * 26)


if __name__ == '__main__':

    main_path = os.path.join(os.path.split(os.path.abspath(sys.argv[0]))[0], '_config')
    cf = configparser.ConfigParser()
    cf.read(''.join((main_path, '\\', 'config.ini')), encoding='utf-8-SIG')
    target_path = cf.get('main', 'target_path').split(',')[0]
    yesterday = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m%d')
    if cf.get('main', 'timing').split(',')[0] == '1':
        target_path = '\\'.join((target_path, yesterday))
    if not os.path.exists(target_path):
        os.makedirs(target_path)
    # if os.path.exists(''.join((target_path, '/LOG_Parser.txt'))):
    #     os.remove(''.join((target_path, '/LOG_Parser.txt')))
    logging.basicConfig(level=logging.INFO,
                        format='',
                        filename=''.join((target_path, '/LOG_Parser.txt')),
                        filemode='a',
                        )
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    logging.getLogger('').addHandler(console)

    # print('main:', os.getpid())
    multiprocessing.freeze_support()
    star_time = time.time()
    # 初始化配置
    config_manager = Main()
    logging.info(time.strftime('%Y/%m/%d %H:%M:%S', time.localtime()))
    # print(time.strftime('%Y/%m/%d %H:%M:%S', time.localtime()))
    parse_type = config_manager.config_main['parse_type']
    # 统计获取到的文件数：
    exit_num = 0
    for temp_i in config_manager.parse_file_list:
        num = 0
        for temp_k in config_manager.parse_file_list[temp_i]:
            num += len(config_manager.parse_file_list[temp_i][temp_k])
        if num != 0:
            logging.info(''.join(('>>> 获取到 ', str(temp_i), ' 文件:', str(num))))
            # print('>>> 获取到 ', temp_i, ' 文件:', num)
            config_manager.all_num[temp_i] = num
            exit_num = 1
    if exit_num == 0:
        logging.info('>>> {0} 没有获取到需处理数据，请检查！'.format(config_manager.config_main['source_path'][0]))
        # print('>>> {0} 没有获取到需处理数据，请检查！'.format(config_manager.config_main['source_path'][0]))
        sys.exit()

    for b in parse_type:
        if b in config_manager.parse_file_list:
            config_manager.run_manager(b.lower())
    logging.info(''.join((time.strftime('%Y/%m/%d %H:%M:%S', time.localtime()))))
    logging.info(''.join(('>>> 历时：', time.strftime('%Y/%m/%d %H:%M:%S', time.gmtime(time.time() - star_time)))))
    logging.info('\n'*2)
    logging.info('='*128)
    logging.info('='*128)